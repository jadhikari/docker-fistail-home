from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import models
from django.db.models import Q, Count, Sum
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.utils import timezone
from datetime import date, timedelta
import calendar
from decimal import Decimal
from django.core.exceptions import ValidationError
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .models import Business, MunicipalShop, Staff, Dependent, Title, Transaction
from .forms import (
    BusinessForm, MunicipalShopForm, StaffForm, DependentForm, 
    BusinessSearchForm, TransactionDetailsSearchForm,
    RevenueForm, ExpenseForm, TitleForm
)
def _has_full_keisan_permissions(user):
    """Return True if user is superuser or has all keisan transaction CRUD + view perms."""
    if not user.is_authenticated:
        return False
    if user.is_superuser:
        return True
    required_perms = [
        # Business
        'keisan.view_business', 'keisan.add_business', 'keisan.change_business', 'keisan.delete_business',
        # MunicipalShop
        'keisan.view_municipalshop', 'keisan.add_municipalshop', 'keisan.change_municipalshop', 'keisan.delete_municipalshop',
        # Staff
        'keisan.view_staff', 'keisan.add_staff', 'keisan.change_staff', 'keisan.delete_staff',
        # Dependent
        'keisan.view_dependent', 'keisan.add_dependent', 'keisan.change_dependent', 'keisan.delete_dependent',
        # Title
        'keisan.view_title', 'keisan.add_title', 'keisan.change_title', 'keisan.delete_title',
        # Transaction
        'keisan.view_transaction', 'keisan.add_transaction', 'keisan.change_transaction', 'keisan.delete_transaction',
    ]
    return user.has_perms(required_perms)






# Helper functions for prorated calculations
def calculate_prorated_amount(monthly_amount, start_date, end_date, period_start, period_end):
    """
    Calculate prorated amount based on actual usage period within a month.
    
    Args:
        monthly_amount: The full monthly amount (salary or rent)
        start_date: When the person/facility started (employment start or rent start)
        end_date: When the person/facility ended (employment end or rent end, None if still active)
        period_start: Start of the calculation period (search period start)
        period_end: End of the calculation period (search period end)
    
    Returns:
        Decimal: Prorated amount for the overlap period
    """
    if not monthly_amount:
        return Decimal('0.00')
    
    # Convert to Decimal for precise calculations
    monthly_amount = Decimal(str(monthly_amount))
    
    # Determine the actual overlap period
    overlap_start = max(start_date, period_start)
    overlap_end = min(end_date or period_end, period_end)
    
    if overlap_start > overlap_end:
        return Decimal('0.00')
    
    # Calculate total days in the month
    month_start = overlap_start.replace(day=1)
    if overlap_start.month == 12:
        month_end = overlap_start.replace(year=overlap_start.year + 1, month=1, day=1) - timedelta(days=1)
    else:
        month_end = overlap_start.replace(month=overlap_start.month + 1, day=1) - timedelta(days=1)
    
    total_days_in_month = month_end.day
    
    # Calculate actual working days
    actual_start = max(overlap_start, month_start)
    actual_end = min(overlap_end, month_end)
    
    if actual_start > actual_end:
        return Decimal('0.00')
    
    # Calculate days worked
    days_worked = (actual_end - actual_start).days + 1
    
    # Calculate prorated amount
    prorated_amount = (monthly_amount * days_worked) / total_days_in_month
    
    return prorated_amount.quantize(Decimal('0.01'))


def calculate_salary_sheet_amount(monthly_amount, start_date, end_date, month_start, month_end):
    """
    Calculate salary amount for salary sheet display.
    Shows prorated salary based on actual working days within the month.
    
    Args:
        monthly_amount: The full monthly salary amount
        start_date: When the staff started employment
        end_date: When the staff ended employment (None if still active)
        month_start: Start of the month being calculated
        month_end: End of the month being calculated
    
    Returns:
        Decimal: Prorated salary amount based on actual working days
    """
    if not monthly_amount:
        return Decimal('0.00')
    
    # Convert to Decimal for precise calculations
    monthly_amount = Decimal(str(monthly_amount))
    
    # Check if staff was employed during this month
    staff_end_date = end_date or month_end
    
    # If staff was not employed during this month at all
    if start_date > month_end or staff_end_date < month_start:
        return Decimal('0.00')
    
    # Calculate the actual working period within this month
    working_start = max(start_date, month_start)
    working_end = min(staff_end_date, month_end)
    
    # Calculate total days in the month
    if month_start.month == 12:
        next_month = month_start.replace(year=month_start.year + 1, month=1, day=1)
    else:
        next_month = month_start.replace(month=month_start.month + 1, day=1)
    
    total_days_in_month = (next_month - timedelta(days=1)).day
    
    # Calculate actual working days
    working_days = (working_end - working_start).days + 1
    
    # Calculate prorated amount
    prorated_amount = (monthly_amount * working_days) / total_days_in_month
    
    return prorated_amount.quantize(Decimal('0.01'))


def get_month_range(start_date, end_date):
    """
    Get list of months between start_date and end_date (inclusive).
    
    Returns:
        List of tuples (year, month) representing each month in the range
    """
    months = []
    current = start_date.replace(day=1)
    end = end_date.replace(day=1)
    
    while current <= end:
        months.append((current.year, current.month))
        if current.month == 12:
            current = current.replace(year=current.year + 1, month=1)
        else:
            current = current.replace(month=current.month + 1)
    
    return months


# Business Views
@login_required(login_url='/accounts/login/')
def business_list(request):
    """List all businesses with search and filtering - Main Dashboard"""
    # Gate access to keisan dashboard
    if not _has_full_keisan_permissions(request.user):
        messages.error(request, 'You do not have permission to access Keisan.')
        return redirect('hostel:dashboard')
    form = BusinessSearchForm(request.GET)
    businesses = Business.objects.all()
    
    if form.is_valid():
        name = form.cleaned_data.get('name')
        business_type = form.cleaned_data.get('business_type')
        owner_name = form.cleaned_data.get('owner_name')
        
        if name:
            businesses = businesses.filter(name__icontains=name)
        if business_type:
            businesses = businesses.filter(business_type__icontains=business_type)
        if owner_name:
            businesses = businesses.filter(owner_name__icontains=owner_name)
    
    # Dashboard statistics
    total_businesses = Business.objects.count()
    total_shops = MunicipalShop.objects.count()
    total_staff = Staff.objects.count()
    total_dependents = Dependent.objects.count()
    
    # Salary statistics
    total_salary = Staff.objects.aggregate(total=Sum('salary'))['total'] or 0
    total_payroll = total_salary
    
    # Pagination
    paginator = Paginator(businesses, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Title management (for modal)
    titles = Title.objects.all().order_by('category', 'mode', 'name')
    title_form = TitleForm()
    # Permissions for keisan management: superuser or has keisan permissions
    can_manage_keisan = _has_full_keisan_permissions(request.user)

    context = {
        'page_obj': page_obj,
        'form': form,
        'total_count': businesses.count(),
        # Dashboard statistics
        'total_businesses': total_businesses,
        'total_shops': total_shops,
        'total_staff': total_staff,
        'total_dependents': total_dependents,
        'total_payroll': total_payroll,
        # Titles
        'titles': titles,
        'title_form': title_form,
        'can_manage_keisan': can_manage_keisan,
    }
    
    return render(request, 'keisan/business_list.html', context)


@login_required(login_url='/accounts/login/')
@require_POST
def title_create(request):
    if not _has_full_keisan_permissions(request.user):
        messages.error(request, 'You do not have permission to add titles.')
        return redirect('keisan:dashboard')
    form = TitleForm(request.POST)
    if form.is_valid():
        title = form.save(commit=False)
        title.created_by = request.user
        title.updated_by = request.user
        try:
            title.full_clean()
            title.save()
            messages.success(request, 'Title created successfully.')
        except ValidationError as e:
            messages.error(request, '; '.join(sum(e.message_dict.values(), [])))
    else:
        # Collect form errors
        errors = []
        for field, field_errors in form.errors.items():
            for err in field_errors:
                errors.append(f"{field}: {err}")
        messages.error(request, ' '.join(errors) or 'Please correct the errors in the form.')
    return redirect('keisan:dashboard')


@login_required(login_url='/accounts/login/')
@require_POST
def title_update(request, pk):
    if not _has_full_keisan_permissions(request.user):
        messages.error(request, 'You do not have permission to update titles.')
        return redirect('keisan:dashboard')
    title = get_object_or_404(Title, pk=pk)
    form = TitleForm(request.POST, instance=title)
    if form.is_valid():
        title = form.save(commit=False)
        title.updated_by = request.user
        try:
            title.full_clean()
            title.save()
            messages.success(request, 'Title updated successfully.')
        except ValidationError as e:
            messages.error(request, '; '.join(sum(e.message_dict.values(), [])))
    else:
        errors = []
        for field, field_errors in form.errors.items():
            for err in field_errors:
                errors.append(f"{field}: {err}")
        messages.error(request, ' '.join(errors) or 'Please correct the errors in the form.')
    return redirect('keisan:dashboard')


@login_required(login_url='/accounts/login/')
@require_POST
def title_delete(request, pk):
    if not _has_full_keisan_permissions(request.user):
        messages.error(request, 'You do not have permission to delete titles.')
        return redirect('keisan:dashboard')
    title = get_object_or_404(Title, pk=pk)
    try:
        title.delete()
        messages.success(request, 'Title deleted successfully.')
    except Exception as e:
        messages.error(request, f'Could not delete title: {e}')
    return redirect('keisan:dashboard')


@login_required(login_url='/accounts/login/')
def business_detail(request, pk):
    """Show business details"""
    business = get_object_or_404(Business, pk=pk)
    shops = business.shops.all()
    
    # Get status filter for staff
    # Only show staff assigned directly to business (not to shops)
    # Staff assigned to both business and shop should only appear under the shop
    staff_status = request.GET.get('staff_status', 'active')
    if staff_status == 'active':
        staff = business.staff.filter(shop__isnull=True, status='Active')
    elif staff_status == 'inactive':
        staff = business.staff.filter(shop__isnull=True, status__in=['Inactive', 'Terminated', 'Resigned'])
    else:
        staff = business.staff.filter(shop__isnull=True)
    
    context = {
        'business': business,
        'shops': shops,
        'staff': staff,
        'staff_status': staff_status,
    }
    
    return render(request, 'keisan/business_detail.html', context)


@login_required(login_url='/accounts/login/')
def business_create(request):
    """Create new business"""
    if request.method == 'POST':
        form = BusinessForm(request.POST)
        if form.is_valid():
            business = form.save(commit=False)
            business.created_by = request.user
            business.save()
            messages.success(request, f'Business "{business.name}" created successfully!')
            return redirect('keisan:business_list')
    else:
        form = BusinessForm()
    
    context = {'form': form, 'title': 'Create New Business'}
    return render(request, 'keisan/business_form.html', context)


@login_required(login_url='/accounts/login/')
def business_edit(request, pk):
    """Edit existing business"""
    business = get_object_or_404(Business, pk=pk)
    
    if request.method == 'POST':
        form = BusinessForm(request.POST, instance=business)
        if form.is_valid():
            business = form.save(commit=False)
            business.updated_by = request.user
            business.save()
            messages.success(request, f'Business "{business.name}" updated successfully!')
            return redirect('keisan:business_list')
    else:
        form = BusinessForm(instance=business)
    
    context = {'form': form, 'business': business, 'title': f'Edit Business: {business.name}'}
    return render(request, 'keisan/business_form.html', context)


# Municipal Shop Views
@login_required(login_url='/accounts/login/')
def shop_detail(request, pk):
    """Show shop details"""
    shop = get_object_or_404(MunicipalShop, pk=pk)
    
    # Get status filter for staff
    staff_status = request.GET.get('staff_status', 'active')
    if staff_status == 'active':
        staff = shop.staff.filter(status='Active')
    elif staff_status == 'inactive':
        staff = shop.staff.filter(status__in=['Inactive', 'Terminated', 'Resigned'])
    else:
        staff = shop.staff.all()
    
    context = {
        'shop': shop,
        'staff': staff,
        'staff_status': staff_status,
    }
    
    return render(request, 'keisan/shop_detail.html', context)


@login_required(login_url='/accounts/login/')
def shop_create(request):
    """Create new municipal shop"""
    # Get business from query parameter if coming from business details
    business_id = request.GET.get('business')
    initial_data = {}
    
    if business_id:
        try:
            business = Business.objects.get(pk=business_id)
            initial_data['business'] = business
        except Business.DoesNotExist:
            messages.error(request, 'Invalid business selected.')
            return redirect('keisan:business_list')
    
    if request.method == 'POST':
        form = MunicipalShopForm(request.POST)
        if form.is_valid():
            shop = form.save(commit=False)
            shop.created_by = request.user
            shop.save()
            messages.success(request, f'Shop "{shop.name}" created successfully!')
            # Redirect to business detail page if business_id is provided, otherwise to shop detail
            if business_id:
                return redirect('keisan:business_detail', pk=business_id)
            else:
                return redirect('keisan:shop_detail', pk=shop.pk)
    else:
        form = MunicipalShopForm(initial=initial_data)
    
    context = {
        'form': form, 
        'title': 'Create New Municipal Shop',
        'business': initial_data.get('business')
    }
    return render(request, 'keisan/shop_form.html', context)


@login_required(login_url='/accounts/login/')
def shop_edit(request, pk):
    """Edit existing municipal shop"""
    shop = get_object_or_404(MunicipalShop, pk=pk)
    
    if request.method == 'POST':
        form = MunicipalShopForm(request.POST, instance=shop)
        if form.is_valid():
            shop = form.save(commit=False)
            shop.updated_by = request.user
            shop.save()
            messages.success(request, f'Shop "{shop.name}" updated successfully!')
            # Redirect to business detail page since shop is always associated with a business
            return redirect('keisan:business_detail', pk=shop.business.pk)
    else:
        form = MunicipalShopForm(instance=shop)
    
    context = {'form': form, 'shop': shop, 'title': f'Edit Shop: {shop.name}'}
    return render(request, 'keisan/shop_form.html', context)


# Staff Views
@login_required(login_url='/accounts/login/')
def staff_detail(request, pk):
    """Show staff details"""
    staff = get_object_or_404(Staff, pk=pk)
    dependents = staff.dependents.all()
    
    context = {
        'staff': staff,
        'dependents': dependents,
    }
    
    return render(request, 'keisan/staff_detail.html', context)


@login_required(login_url='/accounts/login/')
def staff_create(request):
    """Create new staff member"""
    # Get business or shop from query parameters if coming from business/shop details
    business_id = request.GET.get('business')
    shop_id = request.GET.get('shop')
    initial_data = {}
    pre_selected_business = None
    pre_selected_shop = None
    
    if business_id:
        try:
            pre_selected_business = Business.objects.get(pk=business_id)
            initial_data['business'] = pre_selected_business
        except Business.DoesNotExist:
            messages.error(request, 'Invalid business selected.')
            return redirect('keisan:business_list')
    elif shop_id:
        try:
            pre_selected_shop = MunicipalShop.objects.get(pk=shop_id)
            pre_selected_business = pre_selected_shop.business
            initial_data['shop'] = pre_selected_shop
            initial_data['business'] = pre_selected_business
        except MunicipalShop.DoesNotExist:
            messages.error(request, 'Invalid shop selected.')
            return redirect('keisan:business_list')
    
    if request.method == 'POST':
        form = StaffForm(request.POST, pre_selected_business=pre_selected_business, pre_selected_shop=pre_selected_shop)
        if form.is_valid():
            staff = form.save(commit=False)
            staff.created_by = request.user
            
            # Ensure proper business/shop relationship
            if staff.business and staff.shop:
                if staff.shop.business != staff.business:
                    messages.error(request, 'The selected shop does not belong to the selected business.')
                    context = {'form': form, 'business': pre_selected_business, 'shop': pre_selected_shop}
                    return render(request, 'keisan/staff_form.html', context)
            
            staff.save()
            messages.success(request, f'Staff member "{staff.full_name}" created successfully!')
            # Redirect based on whether staff is associated with business or shop
            if shop_id:
                return redirect('keisan:shop_detail', pk=shop_id)
            elif business_id:
                return redirect('keisan:business_detail', pk=business_id)
            else:
                return redirect('keisan:staff_detail', pk=staff.pk)
        else:
            # Form is not valid, show errors
            messages.error(request, 'Please correct the errors below.')
            if form.errors:
                for field, errors in form.errors.items():
                    for error in errors:
                        messages.error(request, f'{field}: {error}')
    else:
        form = StaffForm(initial=initial_data, pre_selected_business=pre_selected_business, pre_selected_shop=pre_selected_shop)
    
    context = {
        'form': form,
        'business': pre_selected_business,
        'shop': pre_selected_shop,
    }
    return render(request, 'keisan/staff_form.html', context)


@login_required(login_url='/accounts/login/')
def staff_edit(request, pk):
    """Edit existing staff member"""
    staff = get_object_or_404(Staff, pk=pk)
    
    # Check if staff is active
    if staff.status != 'Active':
        messages.error(request, f'Cannot edit {staff.full_name} - Staff member is {staff.status}. Only active staff members can be edited.')
        return redirect('keisan:staff_detail', pk=staff.pk)
    
    # Get the current business and shop for pre-selection
    pre_selected_business = staff.business
    pre_selected_shop = staff.shop
    
    if request.method == 'POST':
        form = StaffForm(request.POST, instance=staff, pre_selected_business=pre_selected_business, pre_selected_shop=pre_selected_shop)
        if form.is_valid():
            staff = form.save(commit=False)
            staff.updated_by = request.user
            
            # Ensure proper business/shop relationship
            if staff.business and staff.shop:
                if staff.shop.business != staff.business:
                    messages.error(request, 'The selected shop does not belong to the selected business.')
                    context = {'form': form, 'staff': staff, 'title': f'Edit Staff Member: {staff.full_name}'}
                    return render(request, 'keisan/staff_form.html', context)
            
            staff.save()
            messages.success(request, f'Staff member "{staff.full_name}" updated successfully!')
            # Redirect based on whether staff is associated with business or shop
            if staff.shop:
                return redirect('keisan:shop_detail', pk=staff.shop.pk)
            elif staff.business:
                return redirect('keisan:business_detail', pk=staff.business.pk)
            else:
                return redirect('keisan:staff_detail', pk=staff.pk)
    else:
        form = StaffForm(instance=staff, pre_selected_business=pre_selected_business, pre_selected_shop=pre_selected_shop)
    
    context = {
        'form': form, 
        'staff': staff, 
        'title': f'Edit Staff Member: {staff.full_name}'
    }
    return render(request, 'keisan/staff_form.html', context)


# Dependent Views
@login_required(login_url='/accounts/login/')
def dependent_create(request, staff_id):
    """Create new dependent for a staff member"""
    staff = get_object_or_404(Staff, pk=staff_id)
    
    # Check if staff is active
    if staff.status != 'Active':
        messages.error(request, f'Cannot add dependent for {staff.full_name} - Staff member is {staff.status}. Only active staff members can have dependents.')
        return redirect('keisan:staff_detail', pk=staff_id)

    if request.method == 'POST':
        form = DependentForm(request.POST)
        if form.is_valid():
            dependent = form.save(commit=False)
            dependent.staff = staff
            dependent.created_by = request.user
            dependent.save()
            messages.success(request, f'Dependent "{dependent.full_name}" added successfully!')
            return redirect('keisan:staff_detail', pk=staff_id)
    else:
        form = DependentForm()
    
    context = {
        'form': form,
        'staff': staff,
        'title': f'Add Dependent for {staff.full_name}',
        'dependent': None  # Explicitly set dependent to None for new creation
    }
    
    return render(request, 'keisan/dependent_form.html', context)


@login_required(login_url='/accounts/login/')
def dependent_edit(request, pk):
    """Edit existing dependent"""
    dependent = get_object_or_404(Dependent, pk=pk)
    staff = dependent.staff
    
    # Check if staff is active
    if staff.status != 'Active':
        messages.error(request, f'Cannot edit dependent {dependent.full_name} - Staff member {staff.full_name} is {staff.status}. Only dependents of active staff members can be edited.')
        return redirect('keisan:staff_detail', pk=staff.pk)
    
    if request.method == 'POST':
        form = DependentForm(request.POST, instance=dependent)
        if form.is_valid():
            dependent = form.save(commit=False)
            dependent.updated_by = request.user
            dependent.save()
            messages.success(request, f'Dependent "{dependent.full_name}" updated successfully!')
            return redirect('keisan:staff_detail', pk=staff.pk)
    else:
        form = DependentForm(instance=dependent)
    
    context = {
        'form': form, 
        'dependent': dependent, 
        'staff': staff,
        'title': f'Edit Dependent: {dependent.full_name}'
    }
    return render(request, 'keisan/dependent_form.html', context)


# AJAX Views for dynamic functionality
@login_required(login_url='/accounts/login/')
def get_business_shops(request, business_id):
    """Get shops for a specific business (AJAX)"""
    try:
        business = Business.objects.get(pk=business_id)
        shops = business.shops.all()
        data = [{'id': shop.pk, 'name': shop.name} for shop in shops]
        return JsonResponse({'shops': data})
    except Business.DoesNotExist:
        return JsonResponse({'error': 'Business not found'}, status=404)


@login_required(login_url='/accounts/login/')
def get_staff_by_business(request, business_id):
    """Get staff for a specific business (AJAX)"""
    try:
        business = Business.objects.get(pk=business_id)
        staff = business.staff.all()
        data = [{'id': staff_member.pk, 'name': staff_member.full_name} for staff_member in staff]
        return JsonResponse({'staff': data})
    except Business.DoesNotExist:
        return JsonResponse({'error': 'Business not found'}, status=404)


@login_required(login_url='/accounts/login/')
def get_shops_for_business(request, business_id):
    """AJAX endpoint to get shops for a specific business"""
    try:
        business = get_object_or_404(Business, pk=business_id)
        shops = business.shops.all().values('id', 'name')
        return JsonResponse({'shops': list(shops)})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


# Transaction Views


@login_required(login_url='/accounts/login/')
def transaction_edit(request, pk):
    """Edit existing transaction - redirect to appropriate form based on transaction type"""
    transaction = get_object_or_404(Transaction, pk=pk)
    
    # Redirect to appropriate edit form based on transaction type
    if transaction.transaction_type == 'Revenue':
        return redirect('keisan:revenue_edit', pk=transaction.pk)
    else:  # Expense
        return redirect('keisan:expense_edit', pk=transaction.pk)


@login_required(login_url='/accounts/login/')
def revenue_edit(request, pk):
    """Edit existing revenue transaction"""
    transaction = get_object_or_404(Transaction, pk=pk, transaction_type='Revenue')
    
    # Get the current business and shop for pre-selection
    pre_selected_business = transaction.business
    pre_selected_shop = transaction.shop
    
    if request.method == 'POST':
        form = RevenueForm(request.POST, instance=transaction, pre_selected_business=pre_selected_business, pre_selected_shop=pre_selected_shop)
        if form.is_valid():
            transaction = form.save(commit=False)
            transaction.updated_by = request.user
            transaction.save()
            messages.success(request, f'Revenue updated successfully!')
            # Redirect to appropriate financial summary based on transaction type
            if transaction.shop:
                return redirect('keisan:shop_financial_summary', pk=transaction.shop.pk)
            else:
                return redirect('keisan:business_financial_summary', pk=transaction.business.pk)
    else:
        # Create form with instance data to ensure proper binding
        initial_data = {
            'title': transaction.title.id if transaction.title else None,
            'transaction_mode': transaction.transaction_mode,
            'amount': transaction.amount,
            'memo': transaction.memo,
            'period': f"{transaction.year}-{transaction.month:02d}",
        }
        form = RevenueForm(instance=transaction, initial=initial_data, pre_selected_business=pre_selected_business, pre_selected_shop=pre_selected_shop)
    
    context = {
        'form': form, 
        'transaction': transaction, 
        'title': f'Edit Revenue'
    }
    return render(request, 'keisan/revenue_form.html', context)


@login_required(login_url='/accounts/login/')
def expense_edit(request, pk):
    """Edit existing expense transaction"""
    transaction = get_object_or_404(Transaction, pk=pk, transaction_type='Expense')
    
    # Get the current business and shop for pre-selection
    pre_selected_business = transaction.business
    pre_selected_shop = transaction.shop
    
    if request.method == 'POST':
        form = ExpenseForm(request.POST, instance=transaction, pre_selected_business=pre_selected_business, pre_selected_shop=pre_selected_shop)
        if form.is_valid():
            transaction = form.save(commit=False)
            transaction.updated_by = request.user
            transaction.save()
            messages.success(request, f'Expense updated successfully!')
            # Redirect to appropriate financial summary based on transaction type
            if transaction.shop:
                return redirect('keisan:shop_financial_summary', pk=transaction.shop.pk)
            else:
                return redirect('keisan:business_financial_summary', pk=transaction.business.pk)
    else:
        # Create form with instance data to ensure proper binding
        initial_data = {
            'title': transaction.title.id if transaction.title else None,
            'transaction_mode': transaction.transaction_mode,
            'amount': transaction.amount,
            'memo': transaction.memo,
            'period': f"{transaction.year}-{transaction.month:02d}",
        }
        form = ExpenseForm(instance=transaction, initial=initial_data, pre_selected_business=pre_selected_business, pre_selected_shop=pre_selected_shop)
    
    context = {
        'form': form, 
        'transaction': transaction, 
        'title': f'Edit Expense'
    }
    return render(request, 'keisan/expense_form.html', context)


@login_required(login_url='/accounts/login/')
def transaction_delete(request, pk):
    """Delete transaction"""
    transaction = get_object_or_404(Transaction, pk=pk)
    
    if request.method == 'POST':
        business_pk = transaction.business.pk
        shop_pk = transaction.shop.pk if transaction.shop else None
        transaction.delete()
        messages.success(request, 'Transaction deleted successfully!')
        
        # Redirect to appropriate financial summary based on transaction type
        if shop_pk:
            return redirect('keisan:shop_financial_summary', pk=shop_pk)
        else:
            return redirect('keisan:business_financial_summary', pk=business_pk)
    
    context = {
        'transaction': transaction,
    }
    return render(request, 'keisan/transaction_confirm_delete.html', context)


@login_required(login_url='/accounts/login/')
def business_financial_summary(request, pk):
    """Show financial summary for a specific business"""
    business = get_object_or_404(Business, pk=pk)
    
    # Get filter parameters
    year = request.GET.get('year', timezone.now().year)
    month = request.GET.get('month')
    
    # Base queryset for business transactions
    transactions = business.transactions.filter(year=year)
    if month:
        transactions = transactions.filter(month=month)
    
    # Calculate totals
    revenue_online = transactions.filter(transaction_type='Revenue', transaction_mode='Online').aggregate(total=Sum('amount'))['total'] or 0
    revenue_offline = transactions.filter(transaction_type='Revenue', transaction_mode='Offline').aggregate(total=Sum('amount'))['total'] or 0
    expense_online = transactions.filter(transaction_type='Expense', transaction_mode='Online').aggregate(total=Sum('amount'))['total'] or 0
    expense_offline = transactions.filter(transaction_type='Expense', transaction_mode='Offline').aggregate(total=Sum('amount'))['total'] or 0
    
    total_revenue = revenue_online + revenue_offline
    total_expense = expense_online + expense_offline
    net_profit = total_revenue - total_expense
    
    # Get transactions by shop
    shop_summaries = []
    for shop in business.shops.all():
        shop_transactions = transactions.filter(shop=shop)
        shop_revenue = shop_transactions.filter(transaction_type='Revenue').aggregate(total=Sum('amount'))['total'] or 0
        shop_expense = shop_transactions.filter(transaction_type='Expense').aggregate(total=Sum('amount'))['total'] or 0
        shop_profit = shop_revenue - shop_expense
        
        shop_summaries.append({
            'shop': shop,
            'revenue': shop_revenue,
            'expense': shop_expense,
            'profit': shop_profit,
            'transaction_count': shop_transactions.count()
        })
    
    # Get transaction lists for business and all shops
    business_revenues = transactions.filter(transaction_type='Revenue', shop__isnull=True).order_by('-year', '-month', '-created_at')
    business_expenses = transactions.filter(transaction_type='Expense', shop__isnull=True).order_by('-year', '-month', '-created_at')
    shop_revenues = transactions.filter(transaction_type='Revenue', shop__isnull=False).order_by('-year', '-month', '-created_at')
    shop_expenses = transactions.filter(transaction_type='Expense', shop__isnull=False).order_by('-year', '-month', '-created_at')
    
    # Calculate totals for each category
    business_revenue_total = business_revenues.aggregate(total=Sum('amount'))['total'] or 0
    business_expense_total = business_expenses.aggregate(total=Sum('amount'))['total'] or 0
    shop_revenue_total = shop_revenues.aggregate(total=Sum('amount'))['total'] or 0
    shop_expense_total = shop_expenses.aggregate(total=Sum('amount'))['total'] or 0
    
    context = {
        'business': business,
        'year': year,
        'month': month,
        'revenue_online': revenue_online,
        'revenue_offline': revenue_offline,
        'expense_online': expense_online,
        'expense_offline': expense_offline,
        'total_revenue': total_revenue,
        'total_expense': total_expense,
        'net_profit': net_profit,
        'shop_summaries': shop_summaries,
        'transaction_count': transactions.count(),
        'business_revenues': business_revenues,
        'business_expenses': business_expenses,
        'shop_revenues': shop_revenues,
        'shop_expenses': shop_expenses,
        'business_revenue_total': business_revenue_total,
        'business_expense_total': business_expense_total,
        'shop_revenue_total': shop_revenue_total,
        'shop_expense_total': shop_expense_total,
    }
    
    return render(request, 'keisan/business_financial_summary.html', context)


@login_required(login_url='/accounts/login/')
def shop_financial_summary(request, pk):
    """Show financial summary for a specific shop"""
    shop = get_object_or_404(MunicipalShop, pk=pk)
    
    # Get filter parameters
    year = request.GET.get('year', timezone.now().year)
    month = request.GET.get('month')
    
    # Base queryset for shop transactions
    transactions = shop.transactions.filter(year=year)
    if month:
        transactions = transactions.filter(month=month)
    
    # Calculate totals
    revenue_online = transactions.filter(transaction_type='Revenue', transaction_mode='Online').aggregate(total=Sum('amount'))['total'] or 0
    revenue_offline = transactions.filter(transaction_type='Revenue', transaction_mode='Offline').aggregate(total=Sum('amount'))['total'] or 0
    expense_online = transactions.filter(transaction_type='Expense', transaction_mode='Online').aggregate(total=Sum('amount'))['total'] or 0
    expense_offline = transactions.filter(transaction_type='Expense', transaction_mode='Offline').aggregate(total=Sum('amount'))['total'] or 0
    
    total_revenue = revenue_online + revenue_offline
    total_expense = expense_online + expense_offline
    net_profit = total_revenue - total_expense
    
    # Get transaction lists for this shop only
    shop_revenues = transactions.filter(transaction_type='Revenue').order_by('-year', '-month', '-created_at')
    shop_expenses = transactions.filter(transaction_type='Expense').order_by('-year', '-month', '-created_at')
    
    # Calculate totals for shop transactions
    shop_revenue_total = shop_revenues.aggregate(total=Sum('amount'))['total'] or 0
    shop_expense_total = shop_expenses.aggregate(total=Sum('amount'))['total'] or 0
    
    context = {
        'shop': shop,
        'year': year,
        'month': month,
        'revenue_online': revenue_online,
        'revenue_offline': revenue_offline,
        'expense_online': expense_online,
        'expense_offline': expense_offline,
        'total_revenue': total_revenue,
        'total_expense': total_expense,
        'net_profit': net_profit,
        'transaction_count': transactions.count(),
        'shop_revenues': shop_revenues,
        'shop_expenses': shop_expenses,
        'shop_revenue_total': shop_revenue_total,
        'shop_expense_total': shop_expense_total,
    }
    
    return render(request, 'keisan/shop_financial_summary.html', context)


# Revenue and Expense Views
@login_required(login_url='/accounts/login/')
def revenue_create(request):
    """Create new revenue transaction"""
    # Get business or shop from query parameters if coming from business/shop details
    business_id = request.GET.get('business')
    shop_id = request.GET.get('shop')
    initial_data = {}
    pre_selected_business = None
    pre_selected_shop = None
    
    if business_id:
        try:
            pre_selected_business = Business.objects.get(pk=business_id)
            initial_data['business'] = pre_selected_business
        except Business.DoesNotExist:
            messages.error(request, 'Invalid business selected.')
            return redirect('keisan:business_list')
    elif shop_id:
        try:
            pre_selected_shop = MunicipalShop.objects.get(pk=shop_id)
            pre_selected_business = pre_selected_shop.business
            initial_data['shop'] = pre_selected_shop
            initial_data['business'] = pre_selected_business
        except MunicipalShop.DoesNotExist:
            messages.error(request, 'Invalid shop selected.')
            return redirect('keisan:business_list')
    
    if request.method == 'POST':
        form = RevenueForm(request.POST, pre_selected_business=pre_selected_business, pre_selected_shop=pre_selected_shop)
        if form.is_valid():
            transaction = form.save(commit=False)
            transaction.created_by = request.user
            transaction.save()
            messages.success(request, f'Revenue transaction created successfully!')
            # Redirect based on whether transaction is associated with business or shop
            if shop_id:
                return redirect('keisan:shop_detail', pk=shop_id)
            elif business_id:
                return redirect('keisan:business_detail', pk=business_id)
            else:
                return redirect('keisan:transaction_detail', pk=transaction.pk)
        else:
            # Form is not valid, show errors
            messages.error(request, 'Please correct the errors below.')
            if form.errors:
                for field, errors in form.errors.items():
                    for error in errors:
                        messages.error(request, f'{field}: {error}')
    else:
        form = RevenueForm(initial=initial_data, pre_selected_business=pre_selected_business, pre_selected_shop=pre_selected_shop)
    
    context = {
        'form': form,
        'business': pre_selected_business,
        'shop': pre_selected_shop,
        'title': 'Add Revenue',
        'transaction_type': 'Revenue'
    }
    return render(request, 'keisan/revenue_form.html', context)


@login_required(login_url='/accounts/login/')
def expense_create(request):
    """Create new expense transaction"""
    # Get business or shop from query parameters if coming from business/shop details
    business_id = request.GET.get('business')
    shop_id = request.GET.get('shop')
    initial_data = {}
    pre_selected_business = None
    pre_selected_shop = None
    
    if business_id:
        try:
            pre_selected_business = Business.objects.get(pk=business_id)
            initial_data['business'] = pre_selected_business
        except Business.DoesNotExist:
            messages.error(request, 'Invalid business selected.')
            return redirect('keisan:business_list')
    elif shop_id:
        try:
            pre_selected_shop = MunicipalShop.objects.get(pk=shop_id)
            pre_selected_business = pre_selected_shop.business
            initial_data['shop'] = pre_selected_shop
            initial_data['business'] = pre_selected_business
        except MunicipalShop.DoesNotExist:
            messages.error(request, 'Invalid shop selected.')
            return redirect('keisan:business_list')
    
    if request.method == 'POST':
        form = ExpenseForm(request.POST, pre_selected_business=pre_selected_business, pre_selected_shop=pre_selected_shop)
        if form.is_valid():
            transaction = form.save(commit=False)
            transaction.created_by = request.user
            transaction.save()
            messages.success(request, f'Expense transaction created successfully!')
            # Redirect based on whether transaction is associated with business or shop
            if shop_id:
                return redirect('keisan:shop_detail', pk=shop_id)
            elif business_id:
                return redirect('keisan:business_detail', pk=business_id)
            else:
                return redirect('keisan:transaction_detail', pk=transaction.pk)
        else:
            # Form is not valid, show errors
            messages.error(request, 'Please correct the errors below.')
            if form.errors:
                for field, errors in form.errors.items():
                    for error in errors:
                        messages.error(request, f'{field}: {error}')
    else:
        form = ExpenseForm(initial=initial_data, pre_selected_business=pre_selected_business, pre_selected_shop=pre_selected_shop)
    
    context = {
        'form': form,
        'business': pre_selected_business,
        'shop': pre_selected_shop,
        'title': 'Add Expense',
        'transaction_type': 'Expense'
    }
    return render(request, 'keisan/expense_form.html', context)


@login_required(login_url='/accounts/login/')
def transaction_details(request):
    """Comprehensive transaction details page with search functionality"""
    if not _has_full_keisan_permissions(request.user):
        messages.error(request, 'You do not have permission to access Keisan transactions.')
        return redirect('hostel:dashboard')
    form = TransactionDetailsSearchForm(request.GET)
    context = {
        'form': form,
        'show_results': False,
        'business': None,
        'business_info': None,
        'business_revenue': None,
        'business_expenses': None,
        'business_staff': None,
        'shops_data': None,
    }
    
    # Check if Excel export is requested
    if request.GET.get('format') == 'excel':
        return export_transaction_details_excel(request)
    
    if form.is_valid():
        business = form.cleaned_data.get('business')
        from_period = form.cleaned_data.get('from_period')
        to_period = form.cleaned_data.get('to_period')
        
        # Parse period inputs (format: YYYY-MM)
        from_year = None
        from_month = None
        to_year = None
        to_month = None
        
        if from_period:
            try:
                from_year, from_month = from_period.split('-')
                from_year = int(from_year)
                from_month = int(from_month)
            except (ValueError, AttributeError):
                from_year = None
                from_month = None
        
        if to_period:
            try:
                to_year, to_month = to_period.split('-')
                to_year = int(to_year)
                to_month = int(to_month)
            except (ValueError, AttributeError):
                to_year = None
                to_month = None
        
        if business:
            context['show_results'] = True
            context['business'] = business
            
            # Business Information
            context['business_info'] = {
                'name': business.name,
                'registration_number': business.registration_number,
                'business_type': business.business_type,
                'industry_category': business.industry_category,
                'email': business.email,
                'phone': business.phone,
                'website': business.website,
                'address': business.address,
                'tax_number': business.tax_number,
                'owner_name': business.owner_name,
                'owner_contact_number': business.owner_contact_number,
                'owner_email': business.owner_email,
                'owner_address': business.owner_address,
                'office_rent': business.office_rent,
            }
            
            # Filter transactions by date range
            business_transactions = business.transactions.all()
            if from_year and from_month:
                # Convert string values to integers
                try:
                    from_year_int = int(from_year)
                    from_month_int = int(from_month)
                    business_transactions = business_transactions.filter(
                        models.Q(year__gt=from_year_int) | 
                        (models.Q(year=from_year_int) & models.Q(month__gte=from_month_int))
                    )
                except (ValueError, TypeError):
                    # If conversion fails, use all transactions
                    pass
            
            if to_year and to_month:
                # Convert string values to integers
                try:
                    to_year_int = int(to_year)
                    to_month_int = int(to_month)
                    business_transactions = business_transactions.filter(
                        models.Q(year__lt=to_year_int) | 
                        (models.Q(year=to_year_int) & models.Q(month__lte=to_month_int))
                    )
                except (ValueError, TypeError):
                    # If conversion fails, use all transactions
                    pass
            
            # Business Revenue
            business_revenue = business_transactions.filter(transaction_type='Revenue', shop__isnull=True)
            
            # Organize revenue by month with detailed categorization
            monthly_revenue = {}
            for transaction in business_revenue:
                month_key = f"{transaction.year}-{transaction.month:02d}"
                if month_key not in monthly_revenue:
                    monthly_revenue[month_key] = {
                        'year': transaction.year,
                        'month': transaction.month,
                        'month_name': transaction.get_month_display(),
                        'transactions': [],
                        'total_amount': 0,
                        'online_total': 0,
                        'offline_total': 0,
                        'online_by_title': {},
                        'offline_by_title': {},
                        'online_transactions_by_title': {},
                        'offline_transactions_by_title': {},
                        'rent_amount': 0,
                    }
                
                monthly_revenue[month_key]['transactions'].append(transaction)
                monthly_revenue[month_key]['total_amount'] += transaction.amount
                
                # Categorize by title and group transactions
                title_name = transaction.title.name
                if transaction.transaction_mode == 'Online':
                    monthly_revenue[month_key]['online_total'] += transaction.amount
                    if title_name not in monthly_revenue[month_key]['online_by_title']:
                        monthly_revenue[month_key]['online_by_title'][title_name] = 0
                        monthly_revenue[month_key]['online_transactions_by_title'][title_name] = []
                    monthly_revenue[month_key]['online_by_title'][title_name] += transaction.amount
                    monthly_revenue[month_key]['online_transactions_by_title'][title_name].append(transaction)
                else:
                    monthly_revenue[month_key]['offline_total'] += transaction.amount
                    if title_name not in monthly_revenue[month_key]['offline_by_title']:
                        monthly_revenue[month_key]['offline_by_title'][title_name] = 0
                        monthly_revenue[month_key]['offline_transactions_by_title'][title_name] = []
                    monthly_revenue[month_key]['offline_by_title'][title_name] += transaction.amount
                    monthly_revenue[month_key]['offline_transactions_by_title'][title_name].append(transaction)
            
            # Convert dictionaries to lists for template access
            for month_data in monthly_revenue.values():
                month_data['online_title_list'] = [
                    {'title': title, 'amount': amount, 'transactions': month_data['online_transactions_by_title'][title]}
                    for title, amount in month_data['online_by_title'].items()
                ]
                month_data['offline_title_list'] = [
                    {'title': title, 'amount': amount, 'transactions': month_data['offline_transactions_by_title'][title]}
                    for title, amount in month_data['offline_by_title'].items()
                ]
            
            # Sort monthly revenue by year and month
            sorted_monthly_revenue = sorted(monthly_revenue.values(), key=lambda x: (x['year'], x['month']), reverse=True)
            
            # Create new table format for business revenue
            # Get all unique titles from the period
            all_online_titles = set()
            all_offline_titles = set()
            
            for month_data in monthly_revenue.values():
                all_online_titles.update(month_data['online_by_title'].keys())
                all_offline_titles.update(month_data['offline_by_title'].keys())
            
            # Sort titles alphabetically
            sorted_online_titles = sorted(all_online_titles)
            sorted_offline_titles = sorted(all_offline_titles)
            
            # Create table data for new format
            business_revenue_table_data = []
            for month_data in sorted_monthly_revenue:
                row_data = {
                    'date': f"{month_data['month_name']} {month_data['year']}",
                    'year': month_data['year'],
                    'month': month_data['month'],
                    'online_titles': {},
                    'offline_titles': {},
                    'online_total': month_data['online_total'],
                    'offline_total': month_data['offline_total'],
                    'grand_total': month_data['total_amount']
                }
                
                # Fill online titles (0 for missing titles)
                for title in sorted_online_titles:
                    row_data['online_titles'][title] = month_data['online_by_title'].get(title, 0)
                
                # Fill offline titles (0 for missing titles)
                for title in sorted_offline_titles:
                    row_data['offline_titles'][title] = month_data['offline_by_title'].get(title, 0)
                
                business_revenue_table_data.append(row_data)
            
            context['business_revenue'] = {
                'transactions': business_revenue.order_by('-year', '-month', '-created_at'),
                'monthly_breakdown': sorted_monthly_revenue,
                'table_data': business_revenue_table_data,
                'online_titles': sorted_online_titles,
                'offline_titles': sorted_offline_titles,
                'total_amount': business_revenue.aggregate(total=Sum('amount'))['total'] or 0,
                'online_total': business_revenue.filter(transaction_mode='Online').aggregate(total=Sum('amount'))['total'] or 0,
                'offline_total': business_revenue.filter(transaction_mode='Offline').aggregate(total=Sum('amount'))['total'] or 0,
            }
            
            # Business Staff with Dependents (needed for expense calculations)
            # Only get staff directly assigned to business (not to shops)
            business_staff = business.staff.filter(shop__isnull=True).order_by('full_name')
            staff_data = []
            
            for staff in business_staff:
                # Check if staff was employed during the selected period
                should_display = True
                if from_year and from_month and to_year and to_month:
                    try:
                        from_year_int = int(from_year)
                        from_month_int = int(from_month)
                        to_year_int = int(to_year)
                        to_month_int = int(to_month)
                        
                        search_start_date = date(from_year_int, from_month_int, 1)
                        search_end_date = date(to_year_int, to_month_int, 1)
                        
                        # Staff employment period
                        staff_start_date = staff.start_date
                        staff_end_date = staff.end_date or search_end_date
                        
                        # Check if staff was employed during the search period
                        if staff_end_date < search_start_date:
                            # Staff left before the search period, don't display
                            should_display = False
                        elif staff_start_date > search_end_date:
                            # Staff started after the search period, don't display
                            should_display = False
                    except (ValueError, TypeError):
                        # If conversion fails, display all staff
                        pass
                
                if should_display:
                    staff_data.append({
                        'staff': staff,
                        'dependents': staff.dependents.all(),
                        'calculated_salary': staff.salary,  # Show monthly salary for display
                        'start_date': staff.start_date,
                        'end_date': staff.end_date,
                        'status': staff.status,
                    })
            
            # Business Expenses
            business_expenses = business_transactions.filter(transaction_type='Expense', shop__isnull=True)
            
            # Calculate total salary expenses for the date range
            total_salary_expense = 0
            calculated_salary_expenses = []
            
            for staff in staff_data:
                if from_year and from_month and to_year and to_month:
                    try:
                        from_year_int = int(from_year)
                        from_month_int = int(from_month)
                        to_year_int = int(to_year)
                        to_month_int = int(to_month)
                        
                        # Calculate overlap between staff employment and search period
                        search_start_date = date(from_year_int, from_month_int, 1)
                        search_end_date = date(to_year_int, to_month_int, 1)
                        
                        staff_start_date = staff['staff'].start_date
                        staff_end_date = staff['staff'].end_date or search_end_date
                        
                        overlap_start = max(staff_start_date, search_start_date)
                        overlap_end = min(staff_end_date, search_end_date)
                        
                        if overlap_start <= overlap_end:
                            # Calculate prorated salary for each month in the overlap period
                            months = get_month_range(overlap_start, overlap_end)
                            
                            for year, month in months:
                                # Calculate prorated amount for this specific month
                                month_start = date(year, month, 1)
                                if month == 12:
                                    month_end = date(year + 1, 1, 1) - timedelta(days=1)
                                else:
                                    month_end = date(year, month + 1, 1) - timedelta(days=1)
                                
                                prorated_amount = calculate_prorated_amount(
                                    staff['staff'].salary,
                                    staff_start_date,
                                    staff_end_date,
                                    month_start,
                                    month_end
                                )
                                
                                if prorated_amount > 0:
                                    total_salary_expense += prorated_amount
                                    calculated_salary_expenses.append({
                                        'year': year,
                                        'month': month,
                                        'amount': prorated_amount,
                                        'title_name': f'Salary - {staff["staff"].full_name}',
                                        'transaction_mode': 'Offline',
                                        'memo': f'Prorated salary for {staff["staff"].full_name} ({month_start.strftime("%B %Y")})',
                                        'is_calculated': True
                                    })
                    except (ValueError, TypeError):
                        pass
            
            # Calculate office rent expenses for the date range
            total_office_rent_expense = 0
            calculated_rent_expenses = []
            
            if from_year and from_month and to_year and to_month:
                try:
                    from_year_int = int(from_year)
                    from_month_int = int(from_month)
                    to_year_int = int(to_year)
                    to_month_int = int(to_month)
                    
                    period_start = date(from_year_int, from_month_int, 1)
                    period_end = date(to_year_int, to_month_int, 1)
                    
                    # Calculate prorated office rent for each month in the period
                    months = get_month_range(period_start, period_end)
                    
                    for year, month in months:
                        # Calculate prorated amount for this specific month
                        month_start = date(year, month, 1)
                        if month == 12:
                            month_end = date(year + 1, 1, 1) - timedelta(days=1)
                        else:
                            month_end = date(year, month + 1, 1) - timedelta(days=1)
                        
                        # For office rent, we assume it's for the full month unless business started/ended
                        # We need to check if business has start/end dates (this would need to be added to Business model)
                        # For now, assume full month rent
                        prorated_amount = business.office_rent
                        
                        if prorated_amount > 0:
                            total_office_rent_expense += prorated_amount
                            calculated_rent_expenses.append({
                                'year': year,
                                'month': month,
                                'amount': prorated_amount,
                                'title_name': 'Office Rent',
                                'transaction_mode': 'Offline',
                                'memo': f'Office rent ({month_start.strftime("%B %Y")})',
                                'is_calculated': True
                            })
                except (ValueError, TypeError):
                    pass
            
            # Combine actual transactions with calculated expenses
            all_expenses = list(business_expenses.order_by('-year', '-month', '-created_at'))
            
            # Add calculated expenses
            for expense in calculated_salary_expenses + calculated_rent_expenses:
                all_expenses.append(expense)
            
            # Sort by year, month (handle both Transaction objects and dictionaries)
            def get_sort_key(item):
                if hasattr(item, 'year') and hasattr(item, 'month'):
                    # Transaction object
                    return (item.year, item.month)
                elif isinstance(item, dict) and 'year' in item and 'month' in item:
                    # Dictionary
                    return (item['year'], item['month'])
                else:
                    return (0, 0)
            
            all_expenses.sort(key=get_sort_key, reverse=True)
            
            # Organize expenses by month with detailed categorization
            monthly_expenses = {}
            for expense in all_expenses:
                if hasattr(expense, 'year') and hasattr(expense, 'month'):
                    # Transaction object
                    year = expense.year
                    month = expense.month
                    month_name = expense.get_month_display()
                    transaction_mode = expense.transaction_mode
                    amount = expense.amount
                    title_name = expense.title.name if hasattr(expense, 'title') else expense.get('title_name', 'Unknown')
                elif isinstance(expense, dict) and 'year' in expense and 'month' in expense:
                    # Dictionary
                    year = expense['year']
                    month = expense['month']
                    month_name = date(year, month, 1).strftime('%B')
                    transaction_mode = expense.get('transaction_mode', 'Offline')
                    amount = expense['amount']
                    title_name = expense.get('title_name', 'Unknown')
                else:
                    continue
                
                month_key = f"{year}-{month:02d}"
                if month_key not in monthly_expenses:
                    monthly_expenses[month_key] = {
                        'year': year,
                        'month': month,
                        'month_name': month_name,
                        'transactions': [],
                        'total_amount': 0,
                        'online_total': 0,
                        'offline_total': 0,
                        'online_by_title': {},
                        'offline_by_title': {},
                        'online_transactions_by_title': {},
                        'offline_transactions_by_title': {},
                        'rent_amount': 0,
                        'salary_amount': 0,
                    }
                
                monthly_expenses[month_key]['transactions'].append(expense)
                monthly_expenses[month_key]['total_amount'] += amount
                
                # Categorize by title and group transactions
                if transaction_mode == 'Online':
                    monthly_expenses[month_key]['online_total'] += amount
                    if title_name not in monthly_expenses[month_key]['online_by_title']:
                        monthly_expenses[month_key]['online_by_title'][title_name] = 0
                        monthly_expenses[month_key]['online_transactions_by_title'][title_name] = []
                    monthly_expenses[month_key]['online_by_title'][title_name] += amount
                    monthly_expenses[month_key]['online_transactions_by_title'][title_name].append(expense)
                else:
                    monthly_expenses[month_key]['offline_total'] += amount
                    if title_name not in monthly_expenses[month_key]['offline_by_title']:
                        monthly_expenses[month_key]['offline_by_title'][title_name] = 0
                        monthly_expenses[month_key]['offline_transactions_by_title'][title_name] = []
                    monthly_expenses[month_key]['offline_by_title'][title_name] += amount
                    monthly_expenses[month_key]['offline_transactions_by_title'][title_name].append(expense)
                
                # Check if this is rent or salary
                if 'rent' in title_name.lower():
                    monthly_expenses[month_key]['rent_amount'] += amount
                elif 'salary' in title_name.lower():
                    monthly_expenses[month_key]['salary_amount'] += amount
            
            # Convert dictionaries to lists for template access
            for month_data in monthly_expenses.values():
                month_data['online_title_list'] = [
                    {'title': title, 'amount': amount, 'transactions': month_data['online_transactions_by_title'][title]}
                    for title, amount in month_data['online_by_title'].items()
                ]
                month_data['offline_title_list'] = [
                    {'title': title, 'amount': amount, 'transactions': month_data['offline_transactions_by_title'][title]}
                    for title, amount in month_data['offline_by_title'].items()
                ]
            
            # Sort monthly expenses by year and month
            sorted_monthly_expenses = sorted(monthly_expenses.values(), key=lambda x: (x['year'], x['month']), reverse=True)
            
            # Create new table format for business expenses
            # Get all unique titles from the period (excluding salary)
            all_online_titles = set()
            all_offline_titles = set()
            
            for month_data in monthly_expenses.values():
                # Filter out salary-related titles from online
                online_titles = {title for title in month_data['online_by_title'].keys() if 'salary' not in title.lower()}
                # Filter out salary and any rent from offline
                offline_titles = {title for title in month_data['offline_by_title'].keys() if 'salary' not in title.lower() and 'rent' not in title.lower()}
                
                all_online_titles.update(online_titles)
                all_offline_titles.update(offline_titles)
            
            # Sort titles alphabetically
            sorted_online_titles = sorted(all_online_titles)
            sorted_offline_titles = sorted(all_offline_titles)
            
            # Create table data for new format
            business_expenses_table_data = []
            for month_data in sorted_monthly_expenses:
                # Calculate offline total excluding salary and any rent (same filter as title collection)
                filtered_offline_total = sum(
                    amount for title, amount in month_data['offline_by_title'].items() 
                    if 'salary' not in title.lower() and 'rent' not in title.lower()
                )
                
                row_data = {
                    'date': f"{month_data['month_name']} {month_data['year']}",
                    'year': month_data['year'],
                    'month': month_data['month'],
                    'rent_amount': month_data['rent_amount'],
                    'online_titles': {},
                    'offline_titles': {},
                    'online_total': month_data['online_total'],
                    'offline_total': filtered_offline_total,
                    'grand_total': month_data['online_total'] + filtered_offline_total + month_data['rent_amount']
                }
                
                # Fill online titles (0 for missing titles, exclude salary)
                for title in sorted_online_titles:
                    if 'salary' not in title.lower():
                        row_data['online_titles'][title] = month_data['online_by_title'].get(title, 0)
                
                # Fill offline titles (0 for missing titles, exclude only office rent)
                for title in sorted_offline_titles:
                    if 'office rent' not in title.lower():
                        row_data['offline_titles'][title] = month_data['offline_by_title'].get(title, 0)
                
                business_expenses_table_data.append(row_data)
            
            context['business_expenses'] = {
                'transactions': all_expenses,
                'monthly_breakdown': sorted_monthly_expenses,
                'table_data': business_expenses_table_data,
                'online_titles': sorted_online_titles,
                'offline_titles': sorted_offline_titles,
                'total_amount': (business_expenses.aggregate(total=Sum('amount'))['total'] or 0) + total_salary_expense + total_office_rent_expense,
                'online_total': business_expenses.filter(transaction_mode='Online').aggregate(total=Sum('amount'))['total'] or 0,
                'offline_total': business_expenses.filter(transaction_mode='Offline').aggregate(total=Sum('amount'))['total'] or 0,
                'calculated_salary_total': total_salary_expense,
                'calculated_rent_total': total_office_rent_expense,
            }
            
            # Set business staff data in context
            context['business_staff'] = staff_data
            
            # Create salary sheet data
            salary_sheet = []
            months_in_range = []
            
            if from_year and from_month and to_year and to_month:
                try:
                    from_year_int = int(from_year)
                    from_month_int = int(from_month)
                    to_year_int = int(to_year)
                    to_month_int = int(to_month)
                    
                    # Get all months in the range
                    current_year = from_year_int
                    current_month = from_month_int
                    
                    while (current_year < to_year_int) or (current_year == to_year_int and current_month <= to_month_int):
                        months_in_range.append((current_year, current_month))
                        current_month += 1
                        if current_month > 12:
                            current_month = 1
                            current_year += 1
                    
                    # Create salary data for each staff member
                    for staff_info in staff_data:
                        staff = staff_info['staff']
                        staff_salary_data = {
                            'staff': staff,
                            'employment_type': staff.get_employment_type_display(),
                            'monthly_salaries': {}
                        }
                        
                        # Calculate salary for each month in range
                        for year, month in months_in_range:
                            month_key = f"{year}-{month:02d}"
                            month_name = date(year, month, 1).strftime('%B')
                            
                            # Check if staff was employed during this month
                            month_start = date(year, month, 1)
                            if month == 12:
                                month_end = date(year + 1, 1, 1) - timedelta(days=1)
                            else:
                                month_end = date(year, month + 1, 1) - timedelta(days=1)
                            
                            staff_start_date = staff.start_date
                            staff_end_date = staff.end_date or month_end
                            
                            # Calculate salary for salary sheet (full salary if employed for the month)
                            salary_amount = calculate_salary_sheet_amount(
                                staff.salary,
                                staff_start_date,
                                staff_end_date,
                                month_start,
                                month_end
                            )
                            
                            staff_salary_data['monthly_salaries'][month_key] = {
                                'year': year,
                                'month': month,
                                'month_name': month_name,
                                'amount': salary_amount,
                                'is_active': staff_start_date <= month_end and staff_end_date >= month_start
                            }
                        
                        salary_sheet.append(staff_salary_data)
                        
                except (ValueError, TypeError):
                    # If date conversion fails, create empty salary sheet
                    salary_sheet = []
            
            context['salary_sheet'] = salary_sheet
            context['months_in_range'] = months_in_range
            
            # Shops Data
            shops_data = []
            for shop in business.shops.all():
                # Shop Information
                shop_info = {
                    'name': shop.name,
                    'permit_id': shop.permit_id,
                    'shop_type': shop.shop_type,
                    'address': shop.address,
                    'shop_rent': shop.shop_rent,
                }
                
                # Shop Revenue
                shop_revenue = business_transactions.filter(transaction_type='Revenue', shop=shop)
                
                # Create monthly revenue breakdown for shop
                shop_monthly_revenue = {}
                for transaction in shop_revenue:
                    month_key = f"{transaction.year}-{transaction.month:02d}"
                    if month_key not in shop_monthly_revenue:
                        shop_monthly_revenue[month_key] = {
                            'year': transaction.year,
                            'month': transaction.month,
                            'month_name': transaction.get_month_display(),
                            'total_amount': 0,
                            'online_total': 0,
                            'offline_total': 0,
                            'online_by_title': {},
                            'offline_by_title': {},
                            'online_transactions_by_title': {},
                            'offline_transactions_by_title': {},
                        }
                    
                    month_data = shop_monthly_revenue[month_key]
                    month_data['total_amount'] += transaction.amount
                    
                    if transaction.transaction_mode == 'Online':
                        month_data['online_total'] += transaction.amount
                        title_name = transaction.title.name
                        if title_name not in month_data['online_by_title']:
                            month_data['online_by_title'][title_name] = 0
                            month_data['online_transactions_by_title'][title_name] = []
                        month_data['online_by_title'][title_name] += transaction.amount
                        month_data['online_transactions_by_title'][title_name].append(transaction)
                    else:
                        month_data['offline_total'] += transaction.amount
                        title_name = transaction.title.name
                        if title_name not in month_data['offline_by_title']:
                            month_data['offline_by_title'][title_name] = 0
                            month_data['offline_transactions_by_title'][title_name] = []
                        month_data['offline_by_title'][title_name] += transaction.amount
                        month_data['offline_transactions_by_title'][title_name].append(transaction)
                
                # Convert dictionaries to lists for template access
                for month_data in shop_monthly_revenue.values():
                    month_data['online_title_list'] = [
                        {'title': title, 'amount': amount, 'transactions': month_data['online_transactions_by_title'][title]}
                        for title, amount in month_data['online_by_title'].items()
                    ]
                    month_data['offline_title_list'] = [
                        {'title': title, 'amount': amount, 'transactions': month_data['offline_transactions_by_title'][title]}
                        for title, amount in month_data['offline_by_title'].items()
                    ]
                
                # Create table data for shop revenue (similar to business revenue)
                all_online_titles = set()
                all_offline_titles = set()
                
                for month_data in shop_monthly_revenue.values():
                    all_online_titles.update(month_data['online_by_title'].keys())
                    all_offline_titles.update(month_data['offline_by_title'].keys())
                
                # Sort titles alphabetically
                sorted_online_titles = sorted(all_online_titles)
                sorted_offline_titles = sorted(all_offline_titles)
                
                # Sort monthly revenue by year and month
                sorted_shop_monthly_revenue = sorted(
                    shop_monthly_revenue.values(),
                    key=lambda x: (x['year'], x['month'])
                )
                
                # Create table data for new format
                shop_revenue_table_data = []
                for month_data in sorted_shop_monthly_revenue:
                    row_data = {
                        'date': f"{month_data['month_name']} {month_data['year']}",
                        'year': month_data['year'],
                        'month': month_data['month'],
                        'online_titles': {},
                        'offline_titles': {},
                        'online_total': month_data['online_total'],
                        'offline_total': month_data['offline_total'],
                        'grand_total': month_data['total_amount']
                    }
                    
                    # Fill online titles (0 for missing titles)
                    for title in sorted_online_titles:
                        row_data['online_titles'][title] = month_data['online_by_title'].get(title, 0)
                    
                    # Fill offline titles (0 for missing titles)
                    for title in sorted_offline_titles:
                        row_data['offline_titles'][title] = month_data['offline_by_title'].get(title, 0)
                    
                    shop_revenue_table_data.append(row_data)

                shop_revenue_data = {
                    'transactions': shop_revenue.order_by('-year', '-month', '-created_at'),
                    'total_amount': shop_revenue.aggregate(total=Sum('amount'))['total'] or 0,
                    'online_total': shop_revenue.filter(transaction_mode='Online').aggregate(total=Sum('amount'))['total'] or 0,
                    'offline_total': shop_revenue.filter(transaction_mode='Offline').aggregate(total=Sum('amount'))['total'] or 0,
                    'monthly_breakdown': shop_monthly_revenue,
                    'table_data': shop_revenue_table_data,
                    'online_titles': sorted_online_titles,
                    'offline_titles': sorted_offline_titles,
                }
                
                # Shop Staff with Dependents (needed for expense calculations)
                shop_staff = shop.staff.all().order_by('full_name')
                shop_staff_data = []
                
                for staff in shop_staff:
                    # Check if staff was employed during the selected period
                    should_display = True
                    if from_year and from_month and to_year and to_month:
                        try:
                            from_year_int = int(from_year)
                            from_month_int = int(from_month)
                            to_year_int = int(to_year)
                            to_month_int = int(to_month)
                            
                            search_start_date = date(from_year_int, from_month_int, 1)
                            search_end_date = date(to_year_int, to_month_int, 1)
                            
                            # Staff employment period
                            staff_start_date = staff.start_date
                            staff_end_date = staff.end_date or search_end_date
                            
                            # Check if staff was employed during the search period
                            if staff_end_date < search_start_date:
                                # Staff left before the search period, don't display
                                should_display = False
                            elif staff_start_date > search_end_date:
                                # Staff started after the search period, don't display
                                should_display = False
                        except (ValueError, TypeError):
                            # If conversion fails, display all staff
                            pass
                    
                    if should_display:
                        shop_staff_data.append({
                            'staff': staff,
                            'dependents': staff.dependents.all(),
                            'calculated_salary': staff.salary,  # Show monthly salary for display
                            'start_date': staff.start_date,
                            'end_date': staff.end_date,
                            'status': staff.status,
                        })
                
                # Shop Expenses
                shop_expenses = business_transactions.filter(transaction_type='Expense', shop=shop)
                
                # Calculate total salary expenses for shop staff in the date range
                shop_total_salary_expense = 0
                shop_calculated_salary_expenses = []
                
                for staff in shop_staff_data:
                    if from_year and from_month and to_year and to_month:
                        try:
                            from_year_int = int(from_year)
                            from_month_int = int(from_month)
                            to_year_int = int(to_year)
                            to_month_int = int(to_month)
                            
                            # Calculate overlap between staff employment and search period
                            search_start_date = date(from_year_int, from_month_int, 1)
                            search_end_date = date(to_year_int, to_month_int, 1)
                            
                            staff_start_date = staff['staff'].start_date
                            staff_end_date = staff['staff'].end_date or search_end_date
                            
                            overlap_start = max(staff_start_date, search_start_date)
                            overlap_end = min(staff_end_date, search_end_date)
                            
                            if overlap_start <= overlap_end:
                                # Calculate prorated salary for each month in the overlap period
                                months = get_month_range(overlap_start, overlap_end)
                                
                                for year, month in months:
                                    # Calculate prorated amount for this specific month
                                    month_start = date(year, month, 1)
                                    if month == 12:
                                        month_end = date(year + 1, 1, 1) - timedelta(days=1)
                                    else:
                                        month_end = date(year, month + 1, 1) - timedelta(days=1)
                                    
                                    prorated_amount = calculate_prorated_amount(
                                        staff['staff'].salary,
                                        staff_start_date,
                                        staff_end_date,
                                        month_start,
                                        month_end
                                    )
                                    
                                    if prorated_amount > 0:
                                        shop_total_salary_expense += prorated_amount
                                        shop_calculated_salary_expenses.append({
                                            'year': year,
                                            'month': month,
                                            'amount': prorated_amount,
                                            'title_name': f'Salary - {staff["staff"].full_name}',
                                            'transaction_mode': 'Offline',
                                            'memo': f'Prorated salary for {staff["staff"].full_name} ({month_start.strftime("%B %Y")})',
                                            'is_calculated': True
                                        })
                        except (ValueError, TypeError):
                            pass
                
                # Calculate shop rent expenses for the date range
                shop_total_rent_expense = 0
                shop_calculated_rent_expenses = []
                
                if from_year and from_month and to_year and to_month:
                    try:
                        from_year_int = int(from_year)
                        from_month_int = int(from_month)
                        to_year_int = int(to_year)
                        to_month_int = int(to_month)
                        
                        period_start = date(from_year_int, from_month_int, 1)
                        period_end = date(to_year_int, to_month_int, 1)
                        
                        # Calculate prorated shop rent for each month in the period
                        months = get_month_range(period_start, period_end)
                        
                        for year, month in months:
                            # Calculate prorated amount for this specific month
                            month_start = date(year, month, 1)
                            if month == 12:
                                month_end = date(year + 1, 1, 1) - timedelta(days=1)
                            else:
                                month_end = date(year, month + 1, 1) - timedelta(days=1)
                            
                            # For shop rent, we assume it's for the full month unless shop started/ended
                            # We need to check if shop has start/end dates (this would need to be added to MunicipalShop model)
                            # For now, assume full month rent
                            prorated_amount = shop.shop_rent
                            
                            if prorated_amount > 0:
                                shop_total_rent_expense += prorated_amount
                                shop_calculated_rent_expenses.append({
                                    'year': year,
                                    'month': month,
                                    'amount': prorated_amount,
                                    'title_name': 'Shop Rent',
                                    'transaction_mode': 'Offline',
                                    'memo': f'Shop rent for {shop.name} ({month_start.strftime("%B %Y")})',
                                    'is_calculated': True
                                })
                    except (ValueError, TypeError):
                        pass
                
                # Combine actual transactions with calculated expenses
                shop_all_expenses = list(shop_expenses.order_by('-year', '-month', '-created_at'))
                
                # Add calculated expenses
                for expense in shop_calculated_salary_expenses + shop_calculated_rent_expenses:
                    shop_all_expenses.append(expense)
                
                # Sort by year, month (handle both Transaction objects and dictionaries)
                shop_all_expenses.sort(key=get_sort_key, reverse=True)
                
                # Create monthly expense breakdown for shop
                shop_monthly_expenses = {}
                for expense in shop_all_expenses:
                    # Handle both Transaction objects and calculated expense dictionaries
                    if hasattr(expense, 'year') and hasattr(expense, 'month'):
                        # It's a Transaction object
                        year = expense.year
                        month = expense.month
                        amount = expense.amount
                        title_name = expense.title.name
                        transaction_mode = expense.transaction_mode
                        is_calculated = False
                    else:
                        # It's a calculated expense dictionary
                        year = expense['year']
                        month = expense['month']
                        amount = expense['amount']
                        title_name = expense['title_name']
                        transaction_mode = expense['transaction_mode']
                        is_calculated = expense.get('is_calculated', False)
                    
                    month_key = f"{year}-{month:02d}"
                    if month_key not in shop_monthly_expenses:
                        # Get month name
                        month_name = date(year, month, 1).strftime('%B')
                        shop_monthly_expenses[month_key] = {
                            'year': year,
                            'month': month,
                            'month_name': month_name,
                            'total_amount': 0,
                            'online_total': 0,
                            'offline_total': 0,
                            'rent_amount': 0,
                            'salary_amount': 0,
                            'online_by_title': {},
                            'offline_by_title': {},
                            'online_transactions_by_title': {},
                            'offline_transactions_by_title': {},
                        }
                    
                    month_data = shop_monthly_expenses[month_key]
                    month_data['total_amount'] += amount
                    
                    # Track rent and salary amounts
                    if 'Rent' in title_name:
                        month_data['rent_amount'] += amount
                    elif 'Salary' in title_name:
                        month_data['salary_amount'] += amount
                    
                    if transaction_mode == 'Online':
                        month_data['online_total'] += amount
                        if title_name not in month_data['online_by_title']:
                            month_data['online_by_title'][title_name] = 0
                            month_data['online_transactions_by_title'][title_name] = []
                        month_data['online_by_title'][title_name] += amount
                        month_data['online_transactions_by_title'][title_name].append(expense)
                    else:
                        month_data['offline_total'] += amount
                        if title_name not in month_data['offline_by_title']:
                            month_data['offline_by_title'][title_name] = 0
                            month_data['offline_transactions_by_title'][title_name] = []
                        month_data['offline_by_title'][title_name] += amount
                        month_data['offline_transactions_by_title'][title_name].append(expense)
                
                # Convert dictionaries to lists for template access
                for month_data in shop_monthly_expenses.values():
                    month_data['online_title_list'] = [
                        {'title': title, 'amount': amount, 'transactions': month_data['online_transactions_by_title'][title]}
                        for title, amount in month_data['online_by_title'].items()
                    ]
                    month_data['offline_title_list'] = [
                        {'title': title, 'amount': amount, 'transactions': month_data['offline_transactions_by_title'][title]}
                        for title, amount in month_data['offline_by_title'].items()
                    ]
                
                # Create table data for shop expenses (similar to business expenses)
                all_online_titles = set()
                all_offline_titles = set()
                
                for month_data in shop_monthly_expenses.values():
                    # Filter out salary-related titles from online
                    online_titles = {title for title in month_data['online_by_title'].keys() if 'salary' not in title.lower()}
                    # Filter out salary and shop rent from offline
                    offline_titles = {title for title in month_data['offline_by_title'].keys() if 'salary' not in title.lower() and 'shop rent' not in title.lower()}
                    
                    all_online_titles.update(online_titles)
                    all_offline_titles.update(offline_titles)
                
                # Sort titles alphabetically
                sorted_online_titles = sorted(all_online_titles)
                sorted_offline_titles = sorted(all_offline_titles)
                
                # Sort monthly expenses by year and month
                sorted_shop_monthly_expenses = sorted(
                    shop_monthly_expenses.values(),
                    key=lambda x: (x['year'], x['month'])
                )
                
                # Create table data for new format
                shop_expenses_table_data = []
                for month_data in sorted_shop_monthly_expenses:
                    # Calculate offline total excluding salary and shop rent (same filter as title collection)
                    filtered_offline_total = sum(
                        amount for title, amount in month_data['offline_by_title'].items() 
                        if 'salary' not in title.lower() and 'shop rent' not in title.lower()
                    )
                    
                    row_data = {
                        'date': f"{month_data['month_name']} {month_data['year']}",
                        'year': month_data['year'],
                        'month': month_data['month'],
                        'rent_amount': month_data['rent_amount'],
                        'online_titles': {},
                        'offline_titles': {},
                        'online_total': month_data['online_total'],
                        'offline_total': filtered_offline_total,
                        'grand_total': month_data['online_total'] + filtered_offline_total + month_data['rent_amount']
                    }
                    
                    # Fill online titles (0 for missing titles, exclude salary)
                    for title in sorted_online_titles:
                        if 'salary' not in title.lower():
                            row_data['online_titles'][title] = month_data['online_by_title'].get(title, 0)
                    
                    # Fill offline titles (0 for missing titles, exclude only shop rent and salary)
                    for title in sorted_offline_titles:
                        if ('shop rent' not in title.lower() and 
                            'salary' not in title.lower()):
                            row_data['offline_titles'][title] = month_data['offline_by_title'].get(title, 0)
                    
                    shop_expenses_table_data.append(row_data)

                shop_expenses_data = {
                    'transactions': shop_all_expenses,
                    'total_amount': (shop_expenses.aggregate(total=Sum('amount'))['total'] or 0) + shop_total_salary_expense + shop_total_rent_expense,
                    'online_total': shop_expenses.filter(transaction_mode='Online').aggregate(total=Sum('amount'))['total'] or 0,
                    'offline_total': shop_expenses.filter(transaction_mode='Offline').aggregate(total=Sum('amount'))['total'] or 0,
                    'calculated_salary_total': shop_total_salary_expense,
                    'calculated_rent_total': shop_total_rent_expense,
                    'monthly_breakdown': shop_monthly_expenses,
                    'table_data': shop_expenses_table_data,
                    'online_titles': sorted_online_titles,
                    'offline_titles': sorted_offline_titles,
                }
                
                # Create shop salary sheet data
                shop_salary_sheet = []
                if from_year and from_month and to_year and to_month:
                    try:
                        from_year_int = int(from_year)
                        from_month_int = int(from_month)
                        to_year_int = int(to_year)
                        to_month_int = int(to_month)
                        
                        # Create salary data for each shop staff member
                        for staff_info in shop_staff_data:
                            staff = staff_info['staff']
                            staff_salary_data = {
                                'staff': staff,
                                'employment_type': staff.get_employment_type_display(),
                                'monthly_salaries': {}
                            }
                            
                            # Calculate salary for each month in range
                            for year, month in months_in_range:
                                month_key = f"{year}-{month:02d}"
                                month_name = date(year, month, 1).strftime('%B')
                                
                                # Check if staff was employed during this month
                                month_start = date(year, month, 1)
                                if month == 12:
                                    month_end = date(year + 1, 1, 1) - timedelta(days=1)
                                else:
                                    month_end = date(year, month + 1, 1) - timedelta(days=1)
                                
                                staff_start_date = staff.start_date
                                staff_end_date = staff.end_date or month_end
                                
                                # Calculate salary for salary sheet (full salary if employed for the month)
                                salary_amount = calculate_salary_sheet_amount(
                                    staff.salary,
                                    staff_start_date,
                                    staff_end_date,
                                    month_start,
                                    month_end
                                )
                                
                                staff_salary_data['monthly_salaries'][month_key] = {
                                    'year': year,
                                    'month': month,
                                    'month_name': month_name,
                                    'amount': salary_amount,
                                    'is_active': staff_start_date <= month_end and staff_end_date >= month_start
                                }
                            
                            shop_salary_sheet.append(staff_salary_data)
                            
                    except (ValueError, TypeError):
                        # If date conversion fails, create empty salary sheet
                        shop_salary_sheet = []
                
                shops_data.append({
                    'shop_info': shop_info,
                    'revenue': shop_revenue_data,
                    'expenses': shop_expenses_data,
                    'staff': shop_staff_data,
                    'salary_sheet': shop_salary_sheet,
                })
            
            context['shops_data'] = shops_data
    
    return render(request, 'keisan/transaction_details.html', context)


def export_transaction_details_excel(request):
    """Export transaction details to Excel format matching HTML layout"""
    form = TransactionDetailsSearchForm(request.GET)
    
    if not form.is_valid():
        return HttpResponse("Invalid form data", status=400)
    
    business = form.cleaned_data.get('business')
    from_period = form.cleaned_data.get('from_period')
    to_period = form.cleaned_data.get('to_period')
    
    if not business:
        return HttpResponse("No business selected", status=400)
    
    # Parse period for filtering
    from_year = None
    from_month = None
    to_year = None
    to_month = None
    
    if from_period:
        try:
            from_year, from_month = from_period.split('-')
            from_year = int(from_year)
            from_month = int(from_month)
        except (ValueError, AttributeError):
            pass
    
    if to_period:
        try:
            to_year, to_month = to_period.split('-')
            to_year = int(to_year)
            to_month = int(to_month)
        except (ValueError, AttributeError):
            pass
    
    # Filter transactions by date range (same logic as main view)
    business_transactions = business.transactions.all()
    if from_year and from_month:
        business_transactions = business_transactions.filter(
            models.Q(year__gt=from_year) | 
            (models.Q(year=from_year) & models.Q(month__gte=from_month))
        )
    
    if to_year and to_month:
        business_transactions = business_transactions.filter(
            models.Q(year__lt=to_year) | 
            (models.Q(year=to_year) & models.Q(month__lte=to_month))
        )
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Transaction Details - {business.name}"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    row = 1
    
    # Title
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = f"Transaction Details Report - {business.name}"
    ws[f'A{row}'].font = Font(bold=True, size=16)
    ws[f'A{row}'].alignment = center_alignment
    row += 2
    
    # Period information
    if from_period and to_period:
        ws[f'A{row}'] = f"Period: {from_period} to {to_period}"
        ws[f'A{row}'].font = Font(bold=True)
        row += 2
    
    # Business Information
    ws[f'A{row}'] = "Business Information"
    ws[f'A{row}'].font = Font(bold=True, size=14)
    row += 1
    
    business_info = {
        'Name': business.name,
        'Registration Number': business.registration_number,
        'Business Type': business.business_type,
        'Industry Category': business.industry_category or 'N/A',
        'Email': business.email,
        'Phone': business.phone,
        'Website': business.website or 'N/A',
        'Address': business.address or 'N/A',
        'Tax Number': business.tax_number or 'N/A',
        'Owner Name': business.owner_name,
        'Owner Contact': business.owner_contact_number or 'N/A',
        'Owner Email': business.owner_email or 'N/A',
        'Owner Address': business.owner_address or 'N/A',
        'Office Rent': f"¥{business.office_rent:.2f}",
    }
    
    for key, value in business_info.items():
        ws[f'A{row}'] = key
        ws[f'B{row}'] = value
        row += 1
    
    row += 2
    
    # Staff Salary Sheet (matching HTML format)
    business_staff = business.staff.filter(shop__isnull=True).order_by('full_name')
    if business_staff.exists():
        ws[f'A{row}'] = "Staff Salary Sheet"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 1
        
        # Create months range for salary sheet
        months_in_range = []
        if from_year and from_month and to_year and to_month:
            current_year = from_year
            current_month = from_month
            while (current_year < to_year) or (current_year == to_year and current_month <= to_month):
                months_in_range.append((current_year, current_month))
                current_month += 1
                if current_month > 12:
                    current_month = 1
                    current_year += 1
        
        # Salary sheet headers
        headers = ['Staff Name', 'Employment Type']
        for year, month in months_in_range:
            month_name = calendar.month_name[month][:3]  # Jan, Feb, etc.
            headers.append(f"{month_name} {year}")
        headers.append('Total')
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        row += 1
        
        # Staff data
        for staff in business_staff:
            ws[f'A{row}'] = staff.full_name
            ws[f'B{row}'] = staff.get_employment_type_display()
            
            # Add monthly salary data
            col = 3
            for year, month in months_in_range:
                # Check if staff was active during this month
                staff_start_date = staff.start_date
                staff_end_date = staff.end_date or date.today()
                
                month_start = date(year, month, 1)
                if month == 12:
                    month_end = date(year + 1, 1, 1) - timedelta(days=1)
                else:
                    month_end = date(year, month + 1, 1) - timedelta(days=1)
                
                if staff_start_date <= month_end and staff_end_date >= month_start:
                    ws.cell(row=row, column=col, value=f"¥{staff.salary:.2f}")
                else:
                    ws.cell(row=row, column=col, value="-")
                col += 1
            
            # Total salary
            ws.cell(row=row, column=col, value=f"¥{staff.salary:.2f}")
            
            # Add borders
            for col_num in range(1, len(headers) + 1):
                ws.cell(row=row, column=col_num).border = border
            row += 1
    
    row += 2
    
    # Business Revenue - Monthly Breakdown (matching HTML format)
    business_revenue = business_transactions.filter(transaction_type='Revenue', shop__isnull=True)
    
    if business_revenue.exists():
        ws[f'A{row}'] = "Business Revenue - Monthly Breakdown"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 1
        
        # Revenue summary cards
        total_revenue = business_revenue.aggregate(total=Sum('amount'))['total'] or 0
        online_total = business_revenue.filter(transaction_mode='Online').aggregate(total=Sum('amount'))['total'] or 0
        offline_total = business_revenue.filter(transaction_mode='Offline').aggregate(total=Sum('amount'))['total'] or 0
        
        ws[f'A{row}'] = "Total Revenue:"
        ws[f'B{row}'] = f"¥{total_revenue:.2f}"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = "Online Revenue:"
        ws[f'B{row}'] = f"¥{online_total:.2f}"
        row += 1
        
        ws[f'A{row}'] = "Offline Revenue:"
        ws[f'B{row}'] = f"¥{offline_total:.2f}"
        row += 1
        
        ws[f'A{row}'] = "Transactions Count:"
        ws[f'B{row}'] = business_revenue.count()
        row += 2
        
        # Create detailed revenue table (matching HTML table structure)
        # Get all unique titles from the period
        all_online_titles = set()
        all_offline_titles = set()
        
        monthly_revenue = {}
        for transaction in business_revenue:
            month_key = f"{transaction.year}-{transaction.month:02d}"
            if month_key not in monthly_revenue:
                monthly_revenue[month_key] = {
                    'year': transaction.year,
                    'month': transaction.month,
                    'month_name': transaction.get_month_display(),
                    'online_by_title': {},
                    'offline_by_title': {},
                    'online_total': 0,
                    'offline_total': 0,
                }
            
            title_name = transaction.title.name
            if transaction.transaction_mode == 'Online':
                monthly_revenue[month_key]['online_total'] += transaction.amount
                if title_name not in monthly_revenue[month_key]['online_by_title']:
                    monthly_revenue[month_key]['online_by_title'][title_name] = 0
                monthly_revenue[month_key]['online_by_title'][title_name] += transaction.amount
                all_online_titles.add(title_name)
            else:
                monthly_revenue[month_key]['offline_total'] += transaction.amount
                if title_name not in monthly_revenue[month_key]['offline_by_title']:
                    monthly_revenue[month_key]['offline_by_title'][title_name] = 0
                monthly_revenue[month_key]['offline_by_title'][title_name] += transaction.amount
                all_offline_titles.add(title_name)
        
        # Sort titles alphabetically
        sorted_online_titles = sorted(all_online_titles)
        sorted_offline_titles = sorted(all_offline_titles)
        
        # Create table headers (matching HTML structure)
        ws[f'A{row}'] = "Date"
        col = 2
        
        # Online headers
        for title in sorted_online_titles:
            ws.cell(row=row, column=col, value=title)
            col += 1
        ws.cell(row=row, column=col, value="Total(A)")
        col += 1
        
        # Offline headers
        for title in sorted_offline_titles:
            ws.cell(row=row, column=col, value=title)
            col += 1
        ws.cell(row=row, column=col, value="Total(B)")
        col += 1
        ws.cell(row=row, column=col, value="A+B")
        
        # Style headers
        for col_num in range(1, col + 1):
            cell = ws.cell(row=row, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        row += 1
        
        # Add data rows
        for month_data in sorted(monthly_revenue.values(), key=lambda x: (x['year'], x['month']), reverse=True):
            ws[f'A{row}'] = f"{month_data['month_name']} {month_data['year']}"
            col = 2
            
            # Online data
            for title in sorted_online_titles:
                amount = month_data['online_by_title'].get(title, 0)
                ws.cell(row=row, column=col, value=f"¥{amount:.2f}")
                col += 1
            ws.cell(row=row, column=col, value=f"¥{month_data['online_total']:.2f}")
            col += 1
            
            # Offline data
            for title in sorted_offline_titles:
                amount = month_data['offline_by_title'].get(title, 0)
                ws.cell(row=row, column=col, value=f"¥{amount:.2f}")
                col += 1
            ws.cell(row=row, column=col, value=f"¥{month_data['offline_total']:.2f}")
            col += 1
            
            # Grand total
            grand_total = month_data['online_total'] + month_data['offline_total']
            ws.cell(row=row, column=col, value=f"¥{grand_total:.2f}")
            
            # Add borders
            for col_num in range(1, col + 1):
                ws.cell(row=row, column=col_num).border = border
            row += 1
    
    row += 2
    
    # Business Expenses - Monthly Breakdown (matching HTML format)
    business_expenses = business_transactions.filter(transaction_type='Expense', shop__isnull=True)
    
    if business_expenses.exists():
        ws[f'A{row}'] = "Business Expenses - Monthly Breakdown"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 1
        
        # Expense summary cards
        total_expenses = business_expenses.aggregate(total=Sum('amount'))['total'] or 0
        online_expenses = business_expenses.filter(transaction_mode='Online').aggregate(total=Sum('amount'))['total'] or 0
        offline_expenses = business_expenses.filter(transaction_mode='Offline').aggregate(total=Sum('amount'))['total'] or 0
        
        ws[f'A{row}'] = "Total Expenses:"
        ws[f'B{row}'] = f"¥{total_expenses:.2f}"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = "Online Expenses:"
        ws[f'B{row}'] = f"¥{online_expenses:.2f}"
        row += 1
        
        ws[f'A{row}'] = "Offline Expenses:"
        ws[f'B{row}'] = f"¥{offline_expenses:.2f}"
        row += 1
        
        ws[f'A{row}'] = "Transactions Count:"
        ws[f'B{row}'] = business_expenses.count()
        row += 2
        
        # Create detailed expense table (matching HTML structure)
        # Get all unique titles from the period
        all_online_titles = set()
        all_offline_titles = set()
        
        monthly_expenses = {}
        for transaction in business_expenses:
            month_key = f"{transaction.year}-{transaction.month:02d}"
            if month_key not in monthly_expenses:
                monthly_expenses[month_key] = {
                    'year': transaction.year,
                    'month': transaction.month,
                    'month_name': transaction.get_month_display(),
                    'online_by_title': {},
                    'offline_by_title': {},
                    'online_total': 0,
                    'offline_total': 0,
                    'rent_amount': business.office_rent,  # Add rent for each month
                }
            
            title_name = transaction.title.name
            if transaction.transaction_mode == 'Online':
                monthly_expenses[month_key]['online_total'] += transaction.amount
                if title_name not in monthly_expenses[month_key]['online_by_title']:
                    monthly_expenses[month_key]['online_by_title'][title_name] = 0
                monthly_expenses[month_key]['online_by_title'][title_name] += transaction.amount
                all_online_titles.add(title_name)
            else:
                monthly_expenses[month_key]['offline_total'] += transaction.amount
                if title_name not in monthly_expenses[month_key]['offline_by_title']:
                    monthly_expenses[month_key]['offline_by_title'][title_name] = 0
                monthly_expenses[month_key]['offline_by_title'][title_name] += transaction.amount
                all_offline_titles.add(title_name)
        
        # Sort titles alphabetically
        sorted_online_titles = sorted(all_online_titles)
        sorted_offline_titles = sorted(all_offline_titles)
        
        # Create table headers (matching HTML structure)
        ws[f'A{row}'] = "Date"
        col = 2
        
        # Online headers
        for title in sorted_online_titles:
            ws.cell(row=row, column=col, value=title)
            col += 1
        ws.cell(row=row, column=col, value="Total(A)")
        col += 1
        
        # Offline headers
        for title in sorted_offline_titles:
            ws.cell(row=row, column=col, value=title)
            col += 1
        ws.cell(row=row, column=col, value="Total(B)")
        col += 1
        ws.cell(row=row, column=col, value="Rent")
        col += 1
        ws.cell(row=row, column=col, value="A+B+Rent")
        
        # Style headers
        for col_num in range(1, col + 1):
            cell = ws.cell(row=row, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        row += 1
        
        # Add data rows
        for month_data in sorted(monthly_expenses.values(), key=lambda x: (x['year'], x['month']), reverse=True):
            ws[f'A{row}'] = f"{month_data['month_name']} {month_data['year']}"
            col = 2
            
            # Online data
            for title in sorted_online_titles:
                amount = month_data['online_by_title'].get(title, 0)
                ws.cell(row=row, column=col, value=f"¥{amount:.2f}")
                col += 1
            ws.cell(row=row, column=col, value=f"¥{month_data['online_total']:.2f}")
            col += 1
            
            # Offline data
            for title in sorted_offline_titles:
                amount = month_data['offline_by_title'].get(title, 0)
                ws.cell(row=row, column=col, value=f"¥{amount:.2f}")
                col += 1
            ws.cell(row=row, column=col, value=f"¥{month_data['offline_total']:.2f}")
            col += 1
            
            # Rent
            ws.cell(row=row, column=col, value=f"¥{month_data['rent_amount']:.2f}")
            col += 1
            
            # Grand total
            grand_total = month_data['online_total'] + month_data['offline_total'] + month_data['rent_amount']
            ws.cell(row=row, column=col, value=f"¥{grand_total:.2f}")
            
            # Add borders
            for col_num in range(1, col + 1):
                ws.cell(row=row, column=col_num).border = border
            row += 1
    
    row += 2
    
    # Shops Section (matching HTML format)
    shops = business.shops.all()
    if shops.exists():
        for shop in shops:
            ws[f'A{row}'] = f"Shop: {shop.name}"
            ws[f'A{row}'].font = Font(bold=True, size=14)
            row += 1
            
            # Shop Information
            shop_info = {
                'Name': shop.name,
                'Permit ID': shop.permit_id,
                'Shop Type': shop.shop_type,
                'Address': shop.address or 'N/A',
                'Shop Rent': f"¥{shop.shop_rent:.2f}",
            }
            
            for key, value in shop_info.items():
                ws[f'A{row}'] = key
                ws[f'B{row}'] = value
                row += 1
            
            row += 1
            
            # Shop Revenue
            shop_revenue = business_transactions.filter(transaction_type='Revenue', shop=shop)
            if shop_revenue.exists():
                ws[f'A{row}'] = "Shop Revenue - Detailed Breakdown"
                ws[f'A{row}'].font = Font(bold=True, size=12)
                row += 1
                
                # Revenue summary
                total_revenue = shop_revenue.aggregate(total=Sum('amount'))['total'] or 0
                online_total = shop_revenue.filter(transaction_mode='Online').aggregate(total=Sum('amount'))['total'] or 0
                offline_total = shop_revenue.filter(transaction_mode='Offline').aggregate(total=Sum('amount'))['total'] or 0
                
                ws[f'A{row}'] = "Total Revenue:"
                ws[f'B{row}'] = f"¥{total_revenue:.2f}"
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
                
                ws[f'A{row}'] = "Online Revenue:"
                ws[f'B{row}'] = f"¥{online_total:.2f}"
                row += 1
                
                ws[f'A{row}'] = "Offline Revenue:"
                ws[f'B{row}'] = f"¥{offline_total:.2f}"
                row += 2
                
                # Create detailed revenue table for shop
                all_online_titles = set()
                all_offline_titles = set()
                
                monthly_revenue = {}
                for transaction in shop_revenue:
                    month_key = f"{transaction.year}-{transaction.month:02d}"
                    if month_key not in monthly_revenue:
                        monthly_revenue[month_key] = {
                            'year': transaction.year,
                            'month': transaction.month,
                            'month_name': transaction.get_month_display(),
                            'online_by_title': {},
                            'offline_by_title': {},
                            'online_total': 0,
                            'offline_total': 0,
                        }
                    
                    title_name = transaction.title.name
                    if transaction.transaction_mode == 'Online':
                        monthly_revenue[month_key]['online_total'] += transaction.amount
                        if title_name not in monthly_revenue[month_key]['online_by_title']:
                            monthly_revenue[month_key]['online_by_title'][title_name] = 0
                        monthly_revenue[month_key]['online_by_title'][title_name] += transaction.amount
                        all_online_titles.add(title_name)
                    else:
                        monthly_revenue[month_key]['offline_total'] += transaction.amount
                        if title_name not in monthly_revenue[month_key]['offline_by_title']:
                            monthly_revenue[month_key]['offline_by_title'][title_name] = 0
                        monthly_revenue[month_key]['offline_by_title'][title_name] += transaction.amount
                        all_offline_titles.add(title_name)
                
                # Sort titles alphabetically
                sorted_online_titles = sorted(all_online_titles)
                sorted_offline_titles = sorted(all_offline_titles)
                
                # Create table headers
                ws[f'A{row}'] = "Date"
                col = 2
                
                # Online headers
                for title in sorted_online_titles:
                    ws.cell(row=row, column=col, value=title)
                    col += 1
                ws.cell(row=row, column=col, value="Total(A)")
                col += 1
                
                # Offline headers
                for title in sorted_offline_titles:
                    ws.cell(row=row, column=col, value=title)
                    col += 1
                ws.cell(row=row, column=col, value="Total(B)")
                col += 1
                ws.cell(row=row, column=col, value="A+B")
                
                # Style headers
                for col_num in range(1, col + 1):
                    cell = ws.cell(row=row, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
                    cell.border = border
                row += 1
                
                # Add data rows
                for month_data in sorted(monthly_revenue.values(), key=lambda x: (x['year'], x['month']), reverse=True):
                    ws[f'A{row}'] = f"{month_data['month_name']} {month_data['year']}"
                    col = 2
                    
                    # Online data
                    for title in sorted_online_titles:
                        amount = month_data['online_by_title'].get(title, 0)
                        ws.cell(row=row, column=col, value=f"¥{amount:.2f}")
                        col += 1
                    ws.cell(row=row, column=col, value=f"¥{month_data['online_total']:.2f}")
                    col += 1
                    
                    # Offline data
                    for title in sorted_offline_titles:
                        amount = month_data['offline_by_title'].get(title, 0)
                        ws.cell(row=row, column=col, value=f"¥{amount:.2f}")
                        col += 1
                    ws.cell(row=row, column=col, value=f"¥{month_data['offline_total']:.2f}")
                    col += 1
                    
                    # Grand total
                    grand_total = month_data['online_total'] + month_data['offline_total']
                    ws.cell(row=row, column=col, value=f"¥{grand_total:.2f}")
                    
                    # Add borders
                    for col_num in range(1, col + 1):
                        ws.cell(row=row, column=col_num).border = border
                    row += 1
            
            row += 2
            
            # Shop Expenses
            shop_expenses = business_transactions.filter(transaction_type='Expense', shop=shop)
            if shop_expenses.exists():
                ws[f'A{row}'] = "Shop Expenses - Detailed Breakdown"
                ws[f'A{row}'].font = Font(bold=True, size=12)
                row += 1
                
                # Expense summary
                total_expenses = shop_expenses.aggregate(total=Sum('amount'))['total'] or 0
                online_expenses = shop_expenses.filter(transaction_mode='Online').aggregate(total=Sum('amount'))['total'] or 0
                offline_expenses = shop_expenses.filter(transaction_mode='Offline').aggregate(total=Sum('amount'))['total'] or 0
                
                ws[f'A{row}'] = "Total Expenses:"
                ws[f'B{row}'] = f"¥{total_expenses:.2f}"
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
                
                ws[f'A{row}'] = "Online Expenses:"
                ws[f'B{row}'] = f"¥{online_expenses:.2f}"
                row += 1
                
                ws[f'A{row}'] = "Offline Expenses:"
                ws[f'B{row}'] = f"¥{offline_expenses:.2f}"
                row += 2
                
                # Create detailed expense table for shop (similar to revenue)
                all_online_titles = set()
                all_offline_titles = set()
                
                monthly_expenses = {}
                for transaction in shop_expenses:
                    month_key = f"{transaction.year}-{transaction.month:02d}"
                    if month_key not in monthly_expenses:
                        monthly_expenses[month_key] = {
                            'year': transaction.year,
                            'month': transaction.month,
                            'month_name': transaction.get_month_display(),
                            'online_by_title': {},
                            'offline_by_title': {},
                            'online_total': 0,
                            'offline_total': 0,
                            'rent_amount': shop.shop_rent,
                        }
                    
                    title_name = transaction.title.name
                    if transaction.transaction_mode == 'Online':
                        monthly_expenses[month_key]['online_total'] += transaction.amount
                        if title_name not in monthly_expenses[month_key]['online_by_title']:
                            monthly_expenses[month_key]['online_by_title'][title_name] = 0
                        monthly_expenses[month_key]['online_by_title'][title_name] += transaction.amount
                        all_online_titles.add(title_name)
                    else:
                        monthly_expenses[month_key]['offline_total'] += transaction.amount
                        if title_name not in monthly_expenses[month_key]['offline_by_title']:
                            monthly_expenses[month_key]['offline_by_title'][title_name] = 0
                        monthly_expenses[month_key]['offline_by_title'][title_name] += transaction.amount
                        all_offline_titles.add(title_name)
                
                # Sort titles alphabetically
                sorted_online_titles = sorted(all_online_titles)
                sorted_offline_titles = sorted(all_offline_titles)
                
                # Create table headers
                ws[f'A{row}'] = "Date"
                col = 2
                
                # Online headers
                for title in sorted_online_titles:
                    ws.cell(row=row, column=col, value=title)
                    col += 1
                ws.cell(row=row, column=col, value="Total(A)")
                col += 1
                
                # Offline headers
                for title in sorted_offline_titles:
                    ws.cell(row=row, column=col, value=title)
                    col += 1
                ws.cell(row=row, column=col, value="Total(B)")
                col += 1
                ws.cell(row=row, column=col, value="Rent")
                col += 1
                ws.cell(row=row, column=col, value="A+B+Rent")
                
                # Style headers
                for col_num in range(1, col + 1):
                    cell = ws.cell(row=row, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
                    cell.border = border
                row += 1
                
                # Add data rows
                for month_data in sorted(monthly_expenses.values(), key=lambda x: (x['year'], x['month']), reverse=True):
                    ws[f'A{row}'] = f"{month_data['month_name']} {month_data['year']}"
                    col = 2
                    
                    # Online data
                    for title in sorted_online_titles:
                        amount = month_data['online_by_title'].get(title, 0)
                        ws.cell(row=row, column=col, value=f"¥{amount:.2f}")
                        col += 1
                    ws.cell(row=row, column=col, value=f"¥{month_data['online_total']:.2f}")
                    col += 1
                    
                    # Offline data
                    for title in sorted_offline_titles:
                        amount = month_data['offline_by_title'].get(title, 0)
                        ws.cell(row=row, column=col, value=f"¥{amount:.2f}")
                        col += 1
                    ws.cell(row=row, column=col, value=f"¥{month_data['offline_total']:.2f}")
                    col += 1
                    
                    # Rent
                    ws.cell(row=row, column=col, value=f"¥{month_data['rent_amount']:.2f}")
                    col += 1
                    
                    # Grand total
                    grand_total = month_data['online_total'] + month_data['offline_total'] + month_data['rent_amount']
                    ws.cell(row=row, column=col, value=f"¥{grand_total:.2f}")
                    
                    # Add borders
                    for col_num in range(1, col + 1):
                        ws.cell(row=row, column=col_num).border = border
                    row += 1
            
            row += 2
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="transaction_details_{business.name}_{from_period}_{to_period}.xlsx"'
    
    wb.save(response)
    return response
