from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db.models import Q, Sum, Avg, Count
from django.core.paginator import Paginator
from django.utils import timezone
from django.contrib.auth import get_user_model
from django.http import HttpResponse
from django.db.models import F
from .models import Target, RentalContract
from .forms import TargetForm, TargetAssignmentForm, RentalContractForm
import datetime
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import traceback

User = get_user_model()


def is_superuser(user):
    """Check if user is a superuser"""
    return user.is_authenticated and user.is_superuser


@login_required
@user_passes_test(is_superuser)
def target_management(request):
    """Main target management dashboard for super users"""
    
    # Get current month/year
    today = timezone.now().date()
    current_month = today.month
    current_year = today.year
    
    # Get all targets - ordered by latest period first (most recent year/month on top)
    from django.db.models import F
    targets = Target.objects.select_related(
        'target_to', 'assigned_by'
    ).annotate(
        # Create a computed field for better chronological ordering
        # Formula: year * 100 + month ensures proper chronological order
        # Example: 2024 * 100 + 12 = 202412 (December 2024)
        #         2024 * 100 + 1  = 202401 (January 2024)
        # Ordering by -period_order puts latest periods first
        period_order=F('target_year') * 100 + F('target_month')
    ).order_by('-period_order')
    
    # Filter by month/year if provided
    month_filter = request.GET.get('month')
    year_filter = request.GET.get('year')
    user_filter = request.GET.get('user')
    status_filter = request.GET.get('status')
    
    # Apply filters only if they have actual values
    if month_filter and month_filter.strip():
        targets = targets.filter(target_month=int(month_filter))
    if year_filter and year_filter.strip():
        targets = targets.filter(target_year=int(year_filter))
    if user_filter and user_filter.strip():
        targets = targets.filter(target_to_id=int(user_filter))
    if status_filter and status_filter.strip():
        targets = targets.filter(status=status_filter)
    
    # Pagination
    paginator = Paginator(targets, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get all achievements (rental contracts) for progress calculation
    all_achievements = RentalContract.objects.values(
        'target__target_to', 'target__target_year', 'target__target_month'
    ).annotate(
        total_achieved=Sum('agent_fee') + Sum('ad_fee')
    )
    
    # Create a dictionary for easier lookup by user, year, and month
    achievements_by_user_period = {}
    for achievement in all_achievements:
        key = (achievement['target__target_to'], achievement['target__target_year'], achievement['target__target_month'])
        achievements_by_user_period[key] = achievement['total_achieved']
    
    # Add progress data to each target for easier template access
    for target in page_obj:
        key = (target.target_to.id, target.target_year, target.target_month)
        target.progress_amount = achievements_by_user_period.get(key, 0)
        if target.progress_amount > 0:
            # Convert Decimal to float for calculations
            progress_amount = float(target.progress_amount)
            target_amount = float(target.target_amount)
            target.progress_percentage = (progress_amount / target_amount) * 100
        else:
            target.progress_percentage = 0
    
    # Get filter options (exclude superusers from user filter)
    users = User.objects.filter(is_active=True, is_superuser=False).order_by('email')
    years = Target.objects.values_list('target_year', flat=True).distinct().order_by('-target_year')
    months = [
        (1, 'January'), (2, 'February'), (3, 'March'), (4, 'April'),
        (5, 'May'), (6, 'June'), (7, 'July'), (8, 'August'),
        (9, 'September'), (10, 'October'), (11, 'November'), (12, 'December')
    ]
    
    # Get current month name for display
    current_month_name = dict(months)[current_month]
    
    # Calculate statistics for the filtered targets
    total_targets = targets.count()
    completed_targets = targets.filter(status='completed').count()
    active_targets = targets.filter(status='active').count()
    overdue_targets = targets.filter(status='overdue').count()
    
    # Calculate overall achievement percentage
    overall_achievement_percentage = 0
    if total_targets > 0:
        # Calculate total target amount and total achieved amount
        total_target_amount = sum(float(target.target_amount) for target in targets)
        total_achieved_amount = sum(float(target.progress_amount) for target in page_obj)
        
        if total_target_amount > 0:
            overall_achievement_percentage = (total_achieved_amount / total_target_amount) * 100
    
    context = {
        'page_obj': page_obj,
        'users': users,
        'years': years,
        'months': months,
        'current_month': current_month,
        'current_year': current_year,
        'current_month_name': current_month_name,
        'total_targets': total_targets,
        'completed_targets': completed_targets,
        'active_targets': active_targets,
        'overdue_targets': overdue_targets,
        'overall_achievement_percentage': round(overall_achievement_percentage, 1),
        'achievements_by_user_period': achievements_by_user_period,
        'selected_month': int(month_filter) if month_filter and month_filter.strip() else None,
        'selected_year': int(year_filter) if year_filter and year_filter.strip() else None,
        'selected_user': int(user_filter) if user_filter and user_filter.strip() else None,
        'selected_status': status_filter if status_filter and status_filter.strip() else None,
    }
    
    return render(request, 'targets/target_management.html', context)


@login_required
@user_passes_test(is_superuser)
def export_targets_excel(request):
    """Export targets data to Excel file with progress information"""
    
    # Get current month/year
    today = timezone.now().date()
    current_month = today.month
    current_year = today.year
    
    # Get all targets with related data - ordered by latest period first
    from django.db.models import F
    targets = Target.objects.select_related(
        'target_to', 'assigned_by'
    ).annotate(
        # Create a computed field for better chronological ordering
        # Formula: year * 100 + month ensures proper chronological order
        # Example: 2024 * 100 + 12 = 202412 (December 2024)
        #         2024 * 100 + 1  = 202401 (January 2024)
        # Ordering by -period_order puts latest periods first
        period_order=F('target_year') * 100 + F('target_month')
    ).order_by('-period_order')
    
    # Apply filters if provided
    month_filter = request.GET.get('month')
    year_filter = request.GET.get('year')
    user_filter = request.GET.get('user')
    status_filter = request.GET.get('status')
    
    if month_filter and month_filter.strip():
        targets = targets.filter(target_month=int(month_filter))
    if year_filter and year_filter.strip():
        targets = targets.filter(target_year=int(year_filter))
    if user_filter and user_filter.strip():
        targets = targets.filter(target_to_id=int(user_filter))
    if status_filter and status_filter.strip():
        targets = targets.filter(status=status_filter)
    
    # Get all achievements for progress calculation
    all_achievements = RentalContract.objects.values(
        'target__target_to', 'target__target_year', 'target__target_month'
    ).annotate(
        total_achieved=Sum('agent_fee') + Sum('ad_fee')
    )
    
    # Create a dictionary for easier lookup
    achievements_by_user_period = {}
    for achievement in all_achievements:
        key = (achievement['target__target_to'], achievement['target__target_year'], achievement['target__target_month'])
        achievements_by_user_period[key] = achievement['total_achieved']
    
    # Add progress data to each target
    for target in targets:
        key = (target.target_to.id, target.target_year, target.target_month)
        target.progress_amount = achievements_by_user_period.get(key, 0)
        if target.progress_amount > 0:
            progress_amount = float(target.progress_amount)
            target_amount = float(target.target_amount)
            target.progress_percentage = (progress_amount / target_amount) * 100
        else:
            target.progress_percentage = 0
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Targets Report"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    subheader_font = Font(bold=True, color="FFFFFF")
    subheader_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add title
    ws.merge_cells('A1:H1')
    ws['A1'] = f"Targets Report - Generated on {today.strftime('%B %d, %Y')}"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Add filter information
    filter_info = []
    if month_filter:
        months = {1: 'January', 2: 'February', 3: 'March', 4: 'April', 5: 'May', 6: 'June',
                 7: 'July', 8: 'August', 9: 'September', 10: 'October', 11: 'November', 12: 'December'}
        filter_info.append(f"Month: {months.get(int(month_filter), month_filter)}")
    if year_filter:
        filter_info.append(f"Year: {year_filter}")
    if user_filter:
        user = User.objects.filter(id=user_filter).first()
        if user:
            filter_info.append(f"User: {user.email}")
    if status_filter:
        filter_info.append(f"Status: {status_filter.title()}")
    
    if filter_info:
        ws.merge_cells('A2:H2')
        ws['A2'] = f"Filters: {' | '.join(filter_info)}"
        ws['A2'].font = Font(italic=True, size=12)
        ws['A2'].alignment = Alignment(horizontal="center")
        row_offset = 3
    else:
        row_offset = 2
    
    # Add summary statistics
    total_targets = targets.count()
    completed_targets = targets.filter(status='completed').count()
    active_targets = targets.filter(status='active').count()
    overdue_targets = targets.filter(status='overdue').count()
    
    ws.merge_cells(f'A{row_offset}:H{row_offset}')
    ws[f'A{row_offset}'] = "Summary Statistics"
    ws[f'A{row_offset}'].font = Font(bold=True, size=14)
    ws[f'A{row_offset}'].alignment = Alignment(horizontal="center")
    row_offset += 1
    
    # Summary table
    summary_headers = ['Total Targets', 'Completed', 'Active', 'Overdue', 'Overall Achievement %']
    summary_data = [total_targets, completed_targets, active_targets, overdue_targets]
    
    # Calculate overall achievement
    if total_targets > 0:
        total_target_amount = sum(float(target.target_amount) for target in targets)
        total_achieved_amount = sum(float(target.progress_amount) for target in targets)
        if total_target_amount > 0:
            overall_achievement = (total_achieved_amount / total_target_amount) * 100
        else:
            overall_achievement = 0
    else:
        overall_achievement = 0
    
    summary_data.append(round(overall_achievement, 1))
    
    for col, (header, value) in enumerate(zip(summary_headers, summary_data), 1):
        cell = ws.cell(row=row_offset, column=col)
        cell.value = header
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = header_alignment
        cell.border = border
        
        cell = ws.cell(row=row_offset + 1, column=col)
        cell.value = value
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    
    row_offset += 3
    
    # Add table headers
    headers = [
        'User', 'Email', 'Target Period', 'Target Amount (¥)', 
        'Status', 'Progress %', 'Achieved Amount (¥)', 'Assigned By', 'Created Date'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row_offset, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    row_offset += 1
    
    # Add data rows
    if targets.exists():
        for target in targets:
            ws.cell(row=row_offset, column=1, value=target.target_to.first_name or target.target_to.email)
            ws.cell(row=row_offset, column=2, value=target.target_to.email)
            ws.cell(row=row_offset, column=3, value=target.target_period)
            ws.cell(row=row_offset, column=4, value=float(target.target_amount))
            ws.cell(row=row_offset, column=5, value=target.status.title())
            ws.cell(row=row_offset, column=6, value=round(target.progress_percentage, 1))
            ws.cell(row=row_offset, column=7, value=float(target.progress_amount))
            ws.cell(row=row_offset, column=8, value=target.assigned_by.first_name or target.assigned_by.email)
            ws.cell(row=row_offset, column=9, value=target.created_at.strftime('%Y-%m-%d'))
            
            # Apply borders to all cells in the row
            for col in range(1, 10):
                ws.cell(row=row_offset, column=col).border = border
            
            row_offset += 1
    else:
        # Add a message row if no targets
        ws.merge_cells(f'A{row_offset}:I{row_offset}')
        ws[f'A{row_offset}'] = "No targets found matching the specified criteria."
        ws[f'A{row_offset}'].font = Font(italic=True)
        ws[f'A{row_offset}'].alignment = Alignment(horizontal="center")
        ws[f'A{row_offset}'].border = border
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    # Create a descriptive filename based on filters
    filename_parts = ["targets_report"]
    
    if month_filter:
        months = {1: 'Jan', 2: 'Feb', 3: 'Mar', 4: 'Apr', 5: 'May', 6: 'Jun',
                 7: 'Jul', 8: 'Aug', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dec'}
        filename_parts.append(f"month_{months.get(int(month_filter), month_filter)}")
    
    if year_filter:
        filename_parts.append(f"year_{year_filter}")
    
    if user_filter:
        user = User.objects.filter(id=user_filter).first()
        if user:
            filename_parts.append(f"user_{user.email.split('@')[0]}")
    
    if status_filter:
        filename_parts.append(f"status_{status_filter}")
    
    filename_parts.append(today.strftime('%Y%m%d_%H%M%S'))
    filename = f"{'_'.join(filename_parts)}.xlsx"
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required
@user_passes_test(is_superuser)
def assign_target(request):
    """Assign new target to a user"""
    
    if request.method == 'POST':
        form = TargetAssignmentForm(request.POST)
        if form.is_valid():
            target = form.save(commit=False)
            target.assigned_by = request.user
            target.created_by = request.user
            target.save()
            
            # No need to create rental contract - they are independent
            
            messages.success(
                request, 
                f'Target of ¥{target.target_amount} assigned to {target.target_to.first_name or target.target_to.email} for {target.target_period}'
            )
            return redirect('targets:management')
    else:
        form = TargetAssignmentForm()
    
    return render(request, 'targets/assign_target.html', {'form': form})


@login_required
@user_passes_test(is_superuser)
def edit_target(request, target_id):
    """Edit existing target"""
    
    target = get_object_or_404(Target, id=target_id)
    
    if request.method == 'POST':
        form = TargetForm(request.POST, instance=target)
        if form.is_valid():
            target_instance = form.save(commit=False)
            target_instance.updated_by = request.user
            target_instance.save()
            messages.success(request, 'Target updated successfully!')
            return redirect('targets:management')
    else:
        form = TargetForm(instance=target)
    
    return render(request, 'targets/edit_target.html', {
        'form': form, 
        'target': target
    })


@login_required
@user_passes_test(is_superuser)
def delete_target(request, target_id):
    """Delete target"""
    
    target = get_object_or_404(Target, id=target_id)
    
    if request.method == 'POST':
        user_name = target.target_to.first_name or target.target_to.email
        target.delete()
        messages.success(request, f'Target for {user_name} deleted successfully!')
        return redirect('targets:management')
    
    return render(request, 'targets/delete_target.html', {'target': target})


@login_required
def user_profile(request):
    """User profile page showing current targets"""
    
    user = request.user
    today = timezone.now().date()
    
    # Get user's targets - ordered by latest period first (most recent year/month on top)
    from django.db.models import F
    user_targets = Target.objects.filter(
        target_to=user
    ).select_related('assigned_by').annotate(
        # Create a computed field for better chronological ordering
        # Formula: year * 100 + month ensures proper chronological order
        # Example: 2024 * 100 + 12 = 202412 (December 2024)
        #         2024 * 100 + 1  = 202401 (January 2024)
        # Ordering by -period_order puts latest periods first
        period_order=F('target_year') * 100 + F('target_month')
    ).order_by('-period_order')
    
    # Get all achievements (rental contracts) for this user
    user_achievements = RentalContract.objects.filter(
        target__target_to=user
    ).values('target__target_year', 'target__target_month').annotate(
        total_achieved=Sum('agent_fee') + Sum('ad_fee')
    )
    
    # Create a list for easier template iteration
    achievements_by_period = []
    for achievement in user_achievements:
        achievements_by_period.append({
            'year': achievement['target__target_year'],
            'month': achievement['target__target_month'],
            'total_achieved': achievement['total_achieved']
        })
    
    # Get current month target
    current_target = user_targets.filter(
        target_month=today.month,
        target_year=today.year
    ).first()
    
    # Check if user can add achievements (only if they have an active target)
    can_add_achievements = current_target and current_target.status == 'active'
    
    # Get current month achievements (rental contracts for current month targets)
    current_month_achievements = RentalContract.objects.filter(
        target__target_to=user,
        target__target_month=today.month,
        target__target_year=today.year
    ).order_by('-created_at')
    
    # Calculate current month achievement total
    current_month_achievement_total = sum(
        achievement.agent_fee + achievement.ad_fee for achievement in current_month_achievements
    )
    
    # Filter targets based on search parameters
    month_filter = request.GET.get('month')
    year_filter = request.GET.get('year')
    status_filter = request.GET.get('status')
    
    # Apply filters only if they have actual values
    filtered_targets = user_targets
    if month_filter and month_filter.strip():
        filtered_targets = filtered_targets.filter(target_month=int(month_filter))
    if year_filter and year_filter.strip():
        filtered_targets = filtered_targets.filter(target_year=int(year_filter))
    if status_filter and status_filter.strip():
        filtered_targets = filtered_targets.filter(status=status_filter)
    
    # If no filters applied, show all targets by default
    # (removed current month filter - now shows all targets when no filters)
    
    # Pagination
    paginator = Paginator(filtered_targets, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Calculate statistics for filtered results
    total_targets = user_targets.count()
    completed_targets = user_targets.filter(status='completed').count()
    
    completion_rate = 0
    if total_targets > 0:
        completion_rate = (completed_targets / total_targets) * 100
    
    # Get active and overdue counts for current month
    current_month_targets = user_targets.filter(
        target_month=today.month,
        target_year=today.year
    )
    active_targets = current_month_targets.filter(status='active').count()
    overdue_targets = current_month_targets.filter(status='overdue').count()
    
    # Years and months for filter dropdown
    years = range(today.year - 2, today.year + 2)
    months = [
        (1, 'January'), (2, 'February'), (3, 'March'), (4, 'April'),
        (5, 'May'), (6, 'June'), (7, 'July'), (8, 'August'),
        (9, 'September'), (10, 'October'), (11, 'November'), (12, 'December')
    ]
    
    # Get current month name for display
    current_month_name = dict(months)[today.month]
    
    context = {
        'user': user,
        'current_target': current_target,
        'can_add_achievements': can_add_achievements,
        'current_month_achievements': current_month_achievements,
        'current_month_achievement_total': current_month_achievement_total,
        'achievements_by_period': achievements_by_period,
        'page_obj': page_obj,
        'total_targets': total_targets,
        'completed_targets': completed_targets,
        'active_targets': active_targets,
        'overdue_targets': overdue_targets,
        'completion_rate': completion_rate,
        'years': years,
        'months': months,
        'current_month': today.month,
        'current_year': today.year,
        'current_month_name': current_month_name,
        'selected_month': int(month_filter) if month_filter and month_filter.strip() else None,
        'selected_year': int(year_filter) if year_filter and year_filter.strip() else None,
        'selected_status': status_filter if status_filter and status_filter.strip() else None,
    }
    
    return render(request, 'targets/user_profile.html', context)


@login_required
def achievement_details(request, year, month):
    """View achievement details for a specific year/month"""
    
    user = request.user
    
    # Get achievements (rental contracts) for the specified year/month
    achievements = RentalContract.objects.filter(
        target__target_to=user,
        target__target_year=year,
        target__target_month=month
    ).order_by('-created_at')
    
    # Get month name
    month_names = [
        'January', 'February', 'March', 'April', 'May', 'June',
        'July', 'August', 'September', 'October', 'November', 'December'
    ]
    month_name = month_names[month - 1]
    
    # Calculate total achievement amount
    total_amount = sum(achievement.total_amount for achievement in achievements)
    
    context = {
        'user': user,
        'achievements': achievements,
        'year': year,
        'month': month,
        'month_name': month_name,
        'total_amount': total_amount,
    }
    
    return render(request, 'targets/achievement_details.html', context)


@login_required
def create_rental_contract(request):
    """Create a new rental contract (achievement)"""
    
    user = request.user
    today = timezone.now().date()
    
    # Check if user has an active target for the current month
    current_target = Target.objects.filter(
        target_to=user,
        target_month=today.month,
        target_year=today.year,
        status='active'
    ).first()
    
    # if not current_target:
    #     messages.error(
    #         request, 
    #         'You can only add achievements when you have an active target for the current month. '
    #         'Please contact your supervisor if you need a target assigned.'
    #     )
    #     return redirect('targets:profile')
    
    if request.method == 'POST':
        # print(f"POST data received: {request.POST}")
        # print(f"POST data keys: {list(request.POST.keys())}")
        form = RentalContractForm(request.POST)
        # print(f"Form is valid: {form.is_valid()}")
        # if not form.is_valid():
        #     print(f"Form errors: {form.errors}")
        #     print(f"Form non-field errors: {form.non_field_errors}")
        if form.is_valid():
            try:
                rental_contract = form.save(commit=False)
                rental_contract.created_by = request.user
                rental_contract.updated_by = request.user
                rental_contract.target = current_target  # Set the target automatically
                print(f"Target set to: {current_target}")
                print(f"Rental contract data: {rental_contract.__dict__}")
                
                # Save without calling full_clean to avoid validation issues
                rental_contract.save(force_insert=True)
                
                messages.success(
                    request, 
                    f'Achievement added successfully! Rental contract for {rental_contract.customer_name} '
                    f'with total amount: ¥{rental_contract.total_amount} has been added to your current month target.'
                )
                return redirect('targets:profile')
            except Exception as e:
                messages.error(
                    request, 
                    f'Error creating rental contract: {str(e)}. Please try again.'
                )
                # Log the error for debugging
                # print(f"Error creating rental contract: {str(e)}")
                
                traceback.print_exc()
        else:
            # Log form errors for debugging
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(
                        request, 
                        f'Field "{field}": {error}'
                    )
    else:
        # Create form with current target as initial data
        initial_data = {'target': current_target}
        form = RentalContractForm(initial=initial_data)
        
        # Debug: Check if form can be created
        # print(f"Form created successfully with {len(form.fields)} fields")
        # print(f"Form field names: {list(form.fields.keys())}")
        # print(f"Current target: {current_target}")
    
    context = {
        'form': form,
        'current_target': current_target
    }
    
    return render(request, 'targets/rental_contract_form.html', context)


@login_required
@user_passes_test(is_superuser)
def target_statistics(request):
    """Statistics dashboard for super users"""
    
    today = timezone.now().date()
    current_month = today.month
    current_year = today.year
    
    # Overall statistics (exclude superusers from user count)
    total_users = User.objects.filter(is_active=True, is_superuser=False).count()
    total_targets = Target.objects.count()
    active_targets = Target.objects.filter(status='active').count()
    completed_targets = Target.objects.filter(status='completed').count()
    overdue_targets = Target.objects.filter(status='overdue').count()
    cancelled_targets = Target.objects.filter(status='cancelled').count()
    
    # Calculate completion rate
    completion_rate = 0
    if total_targets > 0:
        completion_rate = (completed_targets / total_targets) * 100
    
    # Current month statistics
    current_month_targets = Target.objects.filter(
        target_month=current_month,
        target_year=current_year
    )
    current_month_total = current_month_targets.count()
    current_month_completed = current_month_targets.filter(status='completed').count()
    current_month_active = current_month_targets.filter(status='active').count()
    current_month_overdue = current_month_targets.filter(status='overdue').count()
    
    # Monthly performance (last 12 months)
    monthly_stats = Target.objects.values(
        'target_year', 'target_month'
    ).annotate(
        total_targets=Count('id'),
        completed_targets=Count('id', filter=Q(status='completed')),
        active_targets=Count('id', filter=Q(status='active')),
        overdue_targets=Count('id', filter=Q(status='overdue')),
        avg_target_amount=Avg('target_amount'),
        total_target_amount=Sum('target_amount')
    ).order_by('-target_year', '-target_month')[:12]
    
    # Calculate monthly completion rates
    for stat in monthly_stats:
        if stat['total_targets'] > 0:
            stat['completion_rate'] = (stat['completed_targets'] / stat['total_targets']) * 100
        else:
            stat['completion_rate'] = 0
    
    # Top performers (users with most completed targets)
    user_stats = Target.objects.values(
        'user__first_name', 'user__email'
    ).annotate(
        total_targets=Count('id'),
        completed_targets=Count('id', filter=Q(status='completed')),
        active_targets=Count('id', filter=Q(status='active')),
        overdue_targets=Count('id', filter=Q(status='overdue')),
        avg_target_amount=Avg('target_amount'),
        total_target_amount=Sum('target_amount')
    ).order_by('-completed_targets', '-total_targets')[:10]
    
    # Calculate user completion rates
    for stat in user_stats:
        if stat['total_targets'] > 0:
            stat['completion_rate'] = (stat['completed_targets'] / stat['total_targets']) * 100
        else:
            stat['completion_rate'] = 0
    
    # Yearly summary
    yearly_stats = Target.objects.values('target_year').annotate(
        total_targets=Count('id'),
        completed_targets=Count('id', filter=Q(status='completed')),
        total_amount=Sum('target_amount'),
        avg_amount=Avg('target_amount')
    ).order_by('-target_year')
    
    for stat in yearly_stats:
        if stat['total_targets'] > 0:
            stat['completion_rate'] = (stat['completed_targets'] / stat['total_targets']) * 100
        else:
            stat['completion_rate'] = 0
    
    context = {
        'total_users': total_users,
        'total_targets': total_targets,
        'active_targets': active_targets,
        'completed_targets': completed_targets,
        'overdue_targets': overdue_targets,
        'cancelled_targets': cancelled_targets,
        'completion_rate': completion_rate,
        'current_month_total': current_month_total,
        'current_month_completed': current_month_completed,
        'current_month_active': current_month_active,
        'current_month_overdue': current_month_overdue,
        'monthly_stats': monthly_stats,
        'user_stats': user_stats,
        'yearly_stats': yearly_stats,
        'current_month': current_month,
        'current_year': current_year,
    }
    
    return render(request, 'targets/statistics.html', context)






def target_achievements(request, target_id):
    """View target achievements for a specific target"""
    
    target = get_object_or_404(Target, id=target_id)
    
    # Get achievements (rental contracts) for this specific target
    target_achievements = RentalContract.objects.filter(
        target=target
    ).order_by('-created_at')
    
    # Calculate progress for this specific target
    progress_amount = target_achievements.aggregate(
        total_achieved=Sum('agent_fee') + Sum('ad_fee')
    )['total_achieved'] or 0
    
    if progress_amount > 0:
        progress_amount = float(progress_amount)
        target_amount = float(target.target_amount)
        progress_percentage = (progress_amount / target_amount) * 100
    else:
        progress_percentage = 0
    
    # Convert Decimal fields to float for template calculations
    for achievement in target_achievements:
        achievement.agent_fee_float = float(achievement.agent_fee)
        achievement.ad_fee_float = float(achievement.ad_fee)
        achievement.total_float = achievement.agent_fee_float + achievement.ad_fee_float
    
    # Get filter options
    years = Target.objects.values_list('target_year', flat=True).distinct().order_by('-target_year')
    months = [
        (1, 'January'), (2, 'February'), (3, 'March'), (4, 'April'),
        (5, 'May'), (6, 'June'), (7, 'July'), (8, 'August'),
        (9, 'September'), (10, 'October'), (11, 'November'), (12, 'December')
    ]
    
    context = {
        'target': target,
        'target_achievements': target_achievements,
        'progress_amount': progress_amount,
        'progress_percentage': round(progress_percentage, 1),
        'years': years,
        'months': months,
    }
    
    return render(request, 'targets/target_achievements.html', context)


@login_required
def export_contracts_excel(request):
    """Export contracts to Excel file"""
    
    # Get all contracts with the same filtering logic as the list view
    contracts = RentalContract.objects.select_related('created_by', 'updated_by', 'target', 'target__target_to').order_by('-created_at')
    
    # Apply filters
    customer_name_filter = request.GET.get('customer_name', '').strip()
    customer_phone_filter = request.GET.get('customer_phone', '').strip()
    year_filter = request.GET.get('year', '').strip()
    month_filter = request.GET.get('month', '').strip()
    created_by_filter = request.GET.get('created_by', '').strip()
    
    if customer_name_filter:
        contracts = contracts.filter(customer_name__icontains=customer_name_filter)
    
    if customer_phone_filter:
        contracts = contracts.filter(customer_number__icontains=customer_phone_filter)
    
    if year_filter:
        contracts = contracts.filter(contract_date__year=int(year_filter))
    
    if month_filter:
        contracts = contracts.filter(contract_date__month=int(month_filter))
    
    if created_by_filter:
        contracts = contracts.filter(created_by_id=int(created_by_filter))
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Rental Contracts"
    
    # Define headers
    headers = [
        'Target User', 'Target Period', 'Customer Name', 'Phone', 'Contract Date', 'Contract Type', 'People',
        'Building Address', 'Support Phone', 'Emergency Contact', 'Emergency Phone',
        'Cancellation Notice', 'Cancellation Period', 'Cancellation Charge',
        'Deposit Fee', 'Cleaning Charge', 'Renew Fee', 'Rent Payment Date',
        'Management Company', 'Management Phone', 'Memo',
        'Created Date', 'Created By', 'Updated Date', 'Updated By'
    ]
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    # Add data rows
    for row, contract in enumerate(contracts, 2):
        # Target information
        target_user = contract.target.target_to.get_full_name() if contract.target.target_to.get_full_name() else contract.target.target_to.email
        ws.cell(row=row, column=1, value=target_user)
        ws.cell(row=row, column=2, value=f"{contract.target.target_month}/{contract.target.target_year}")
        
        # Contract information
        ws.cell(row=row, column=3, value=contract.customer_name)
        ws.cell(row=row, column=4, value=contract.customer_number)
        ws.cell(row=row, column=5, value=contract.contract_date)
        ws.cell(row=row, column=6, value=contract.get_contract_type_display())
        ws.cell(row=row, column=7, value=contract.living_num_people)
        ws.cell(row=row, column=8, value=contract.building_address)
        ws.cell(row=row, column=9, value=contract.support_phone)
        ws.cell(row=row, column=10, value=contract.emergency_contact_person)
        ws.cell(row=row, column=11, value=contract.emergency_phone)
        ws.cell(row=row, column=12, value=contract.cancellation_notice_period)
        ws.cell(row=row, column=13, value=contract.cancellation_period)
        ws.cell(row=row, column=14, value=contract.cancellation_charge)
        ws.cell(row=row, column=15, value=contract.deposit_fee)
        ws.cell(row=row, column=16, value="Yes" if contract.cleaning_charge else "No")
        ws.cell(row=row, column=17, value=contract.renew_fee)
        ws.cell(row=row, column=18, value=contract.rent_payment_date)
        
        # Management company information
        ws.cell(row=row, column=19, value=contract.management_company_name or "")
        ws.cell(row=row, column=20, value=contract.management_company_phone_number or "")
        ws.cell(row=row, column=21, value=contract.memo or "")
        
        # Convert timezone-aware datetime to timezone-naive for Excel
        created_at = contract.created_at
        if created_at and created_at.tzinfo:
            created_at = created_at.replace(tzinfo=None)
        ws.cell(row=row, column=22, value=created_at)
        
        # Safely get user display name
        created_by_name = ""
        if hasattr(contract.created_by, 'get_full_name'):
            created_by_name = contract.created_by.get_full_name()
        if not created_by_name:
            created_by_name = contract.created_by.email
        ws.cell(row=row, column=23, value=created_by_name)
        
        # Convert timezone-aware datetime to timezone-naive for Excel
        updated_at = contract.updated_at
        if updated_at and updated_at.tzinfo:
            updated_at = updated_at.replace(tzinfo=None)
        ws.cell(row=row, column=24, value=updated_at)
        
        # Safely get updated by user display name
        updated_by_name = ""
        if contract.updated_by:
            if hasattr(contract.updated_by, 'get_full_name'):
                updated_by_name = contract.updated_by.get_full_name()
            if not updated_by_name:
                updated_by_name = contract.updated_by.email
        ws.cell(row=row, column=25, value=updated_by_name)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=rental_contracts_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb.save(response)
    return response


@login_required
def contracts_list(request):
    """Display all rental contracts with filtering options"""
    
    # Get all contracts ordered by creation date (newest first)
    contracts = RentalContract.objects.select_related('created_by', 'updated_by', 'target', 'target__target_to').order_by('-created_at')
    
    # Apply filters
    customer_name_filter = request.GET.get('customer_name', '').strip()
    customer_phone_filter = request.GET.get('customer_phone', '').strip()
    year_filter = request.GET.get('year', '').strip()
    month_filter = request.GET.get('month', '').strip()
    created_by_filter = request.GET.get('created_by', '').strip()
    
    if customer_name_filter:
        contracts = contracts.filter(customer_name__icontains=customer_name_filter)
    
    if customer_phone_filter:
        contracts = contracts.filter(customer_number__icontains=customer_phone_filter)
    
    if year_filter:
        contracts = contracts.filter(contract_date__year=int(year_filter))
    
    if month_filter:
        contracts = contracts.filter(contract_date__month=int(month_filter))
    
    if created_by_filter:
        contracts = contracts.filter(created_by_id=int(created_by_filter))
    
    # Pagination
    paginator = Paginator(contracts, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get filter options
    years = RentalContract.objects.dates('contract_date', 'year').distinct()
    months = [
        (1, 'January'), (2, 'February'), (3, 'March'), (4, 'April'),
        (5, 'May'), (6, 'June'), (7, 'July'), (8, 'August'),
        (9, 'September'), (10, 'October'), (11, 'November'), (12, 'December')
    ]
    
    # Get all regular users (excluding superusers) for the created_by filter
    users = User.objects.filter(is_active=True, is_superuser=False).order_by('first_name', 'last_name', 'email')
    
    # Calculate total contracts for filtered results
    total_contracts = contracts.count()
    
    context = {
        'contracts': page_obj,
        'total_contracts': total_contracts,
        'years': years,
        'months': months,
        'users': users,
        'filters': {
            'customer_name': customer_name_filter,
            'customer_phone': customer_phone_filter,
            'year': year_filter,
            'month': month_filter,
            'created_by': created_by_filter,
        }
    }
    
    return render(request, 'targets/contracts_list.html', context)


