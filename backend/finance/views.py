from django.shortcuts import render, get_object_or_404,redirect
from .models import HostelRevenue, HostelExpense, UtilityExpense
from hostel.models import Bed
from decimal import Decimal, InvalidOperation
from django.contrib import messages
from .utils import send_revenue_email
from django.contrib.auth.decorators import login_required, permission_required
from django.utils import timezone
from django.db.models import Q, Sum
import openpyxl #type: ignore
from django.http import HttpResponse, JsonResponse
from django.core.paginator import Paginator
from django.views.decorators.csrf import csrf_exempt
import json

from .finance_helpers.rent_defaulters import get_rent_defaulters
from datetime import date, datetime
from .forms import HostelExpenseForm, UtilityExpenseForm
from django.contrib.auth.decorators import login_required, user_passes_test

def export_revenues_to_excel(queryset, record_type='rent'):
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Set worksheet title based on record type
    if record_type == 'registration':
        ws.title = "Registration Revenues"
        headers = [
            'Customer', 'Hostel', 'Unit', 'Bed',
            'Revenue Year', 'Revenue Month',
            'Initial Fee', 'I. F. Discount (%)', 'I. F. After Discount',
            'Deposit', 'D. Discount (%)', 'D. After Discount',
            'Total Amount', 'Created At', 'Created By'
        ]
    else:  # rent records
        ws.title = "Rent Revenues"
        headers = [
            'Customer', 'Hostel', 'Unit', 'Bed',
            'Revenue Year', 'Revenue Month',
            'Internet', 'Utilities', 'Rent', 'Rent Discount (%)', 'Rent After Discount',
            'Total Amount', 'Payment Type', 'Collected Amount', 'Prepaid/Postpaid Amount',
            'Created At', 'Created By'
        ]
    
    ws.append(headers)

    for rev in queryset:
        customer = getattr(rev, 'customer', None)
        bed = getattr(customer, 'bed_assignment', None) if customer else None
        unit = getattr(bed, 'unit', None) if bed else None
        hostel = getattr(unit, 'hostel', None) if unit else None
        
        # Get creator info
        created_by = rev.created_by
        created_by_name = ''
        if created_by:
            if hasattr(created_by, 'first_name') and created_by.first_name:
                created_by_name = created_by.first_name
            elif hasattr(created_by, 'email'):
                created_by_name = created_by.email
            else:
                created_by_name = str(created_by)
        
        created_at = rev.created_at.strftime('%Y-%m-%d %H:%M') if rev.created_at else ''

        if record_type == 'registration':
            # Registration fee export
            ws.append([
                customer.name if customer else '',
                hostel.name if hostel else '',
                unit.room_num if unit else '',
                bed.bed_num if bed else '',
                rev.year,
                rev.month,
                rev.initial_fee or '',
                rev.initial_fee_discount_percent or '',
                rev.initial_fee_after_discount or '',
                rev.deposit or '',
                rev.deposit_discount_percent or '',
                rev.deposit_after_discount or '',
                rev.total_amount or '',
                created_at,
                created_by_name,
            ])
        else:
            # Rent export
            payment_type = ''
            if rev.payment_type == 'prepaid':
                payment_type = 'Prepaid'
            elif rev.payment_type == 'postpaid':
                payment_type = 'Postpaid'
            else:
                payment_type = 'Normal'
            
            prepaid_postpaid_amount = ''
            if rev.payment_type and rev.prepaid_amount:
                if rev.payment_type == 'prepaid':
                    prepaid_postpaid_amount = f'+{rev.prepaid_amount}'
                elif rev.payment_type == 'postpaid':
                    prepaid_postpaid_amount = f'-{rev.prepaid_amount}'
            
            ws.append([
                customer.name if customer else '',
                hostel.name if hostel else '',
                unit.room_num if unit else '',
                bed.bed_num if bed else '',
                rev.year,
                rev.month,
                rev.internet or '',
                rev.utilities or '',
                rev.rent or '',
                rev.rent_discount_percent or '',
                rev.rent_after_discount or '',
                rev.total_amount or '',
                payment_type,
                rev.collected_amount or '',
                prepaid_postpaid_amount,
                created_at,
                created_by_name,
            ])

    # Set filename based on record type
    filename = f"{record_type}_revenues.xlsx"
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required(login_url='/accounts/login/')
def revenues(request):
    name = request.GET.get('name')
    year = request.GET.get('year')
    month = request.GET.get('month')
    hostel = request.GET.get('hostel')
    record_type = request.GET.get('record_type', 'rent')  # Default to 'rent'

    today = timezone.now()
    default_year = today.year
    default_month = today.month

    query = Q()

    if name:
        query &= Q(customer__name__icontains=name)

    if hostel:
        query &= Q(customer__bed_assignment__unit__hostel__name__icontains=hostel)

    # Default filter by creation date instead of revenue year/month
    try:
        selected_year = int(year)
    except (ValueError, TypeError):
        selected_year = default_year

    try:
        selected_month = int(month)
    except (ValueError, TypeError):
        selected_month = default_month

    # Filter by creation date instead of revenue date
    query &= Q(created_at__year=selected_year, created_at__month=selected_month)

    # Filter by record type (rent or registration)
    if record_type == 'registration':
        query &= Q(title='registration_fee')
    else:  # Default to 'rent'
        query &= Q(title='rent')

    # Get all revenues with the base query, ordered by creation date (newest first)
    all_revenues = HostelRevenue.objects.select_related('customer', 'created_by').filter(query).order_by('-created_at')
    
    # Separate registration and rent records based on filter
    if record_type == 'registration':
        registration_revenues = all_revenues
        rent_revenues = all_revenues.none()  # Empty queryset
    else:  # record_type == 'rent' (default)
        registration_revenues = all_revenues.none()  # Empty queryset
        rent_revenues = all_revenues

    # Calculate totals
    registration_total = registration_revenues.aggregate(
        total=Sum('total_amount')
    )['total'] or Decimal('0')
    
    rent_collection_total = rent_revenues.aggregate(
        total=Sum('collected_amount')
    )['total'] or Decimal('0')
    
    rent_total_amount = rent_revenues.aggregate(
        total=Sum('total_amount')
    )['total'] or Decimal('0')

    # No pagination for registration revenues - display all records
    registration_page_obj = registration_revenues

    # Pagination for rent revenues (limit 20)
    rent_page = request.GET.get('rent_page', 1)
    rent_paginator = Paginator(rent_revenues, 20)
    rent_page_obj = rent_paginator.get_page(rent_page)

    # Calculate page totals
    registration_page_total = registration_revenues.aggregate(
        total=Sum('total_amount')
    )['total'] or Decimal('0')
    
    rent_page_total_amount = rent_page_obj.object_list.aggregate(
        total=Sum('total_amount')
    )['total'] or Decimal('0')
    
    rent_page_collection_total = rent_page_obj.object_list.aggregate(
        total=Sum('collected_amount')
    )['total'] or Decimal('0')

    # ✅ Only allow download if there are results
    if request.GET.get('download') == 'excel':
        if all_revenues.exists():
            return export_revenues_to_excel(all_revenues, record_type)
        else:
            messages.warning(request, "No data available to export.")

    if request.GET and not any([name, year, month, hostel]):
        messages.warning(request, "No filter parameters provided.")

    year_choices = HostelRevenue.objects.values_list('year', flat=True).distinct().order_by('-year')

    # Get all hostels for the filter dropdown
    from hostel.models import Hostel
    all_hostels = Hostel.objects.all().order_by('name')

    return render(request, 'finance/revenues_dashboard.html', {
        'registration_revenues': registration_page_obj,
        'rent_revenues': rent_page_obj,
        'registration_total': registration_total,
        'registration_page_total': registration_page_total,
        'rent_collection_total': rent_collection_total,
        'rent_total_amount': rent_total_amount,
        'rent_page_total_amount': rent_page_total_amount,
        'rent_page_collection_total': rent_page_collection_total,
        'name': name,
        'selected_year': selected_year,
        'selected_month': selected_month,
        'selected_hostel': hostel,
        'selected_record_type': record_type,
        'year_choices': year_choices,
        'month_choices': [(i, i) for i in range(1, 13)],
        'all_hostels': all_hostels,
    })

@login_required(login_url='/accounts/login/')
def revenue_detail(request, pk):
    revenue = get_object_or_404(HostelRevenue, pk=pk)
    return render(request, 'finance/revenue_detail.html', {'revenue': revenue})


def get_previous_prepaid_amount(customer, year, month):
    """
    Get the prepaid amount from the previous month's rent payment.
    Returns the prepaid amount if the previous month had a prepaid payment.
    """
    # Calculate previous month
    prev_month = month - 1
    prev_year = year
    if prev_month <= 0:
        prev_month = 12
        prev_year = year - 1
    
    try:
        prev_revenue = HostelRevenue.objects.get(
            title='rent',
            customer=customer,
            year=prev_year,
            month=prev_month
        )
        # Return prepaid amount if it exists and payment type is prepaid
        if prev_revenue.payment_type == 'prepaid' and prev_revenue.prepaid_amount:
            return prev_revenue.prepaid_amount
        return Decimal('0')
    except HostelRevenue.DoesNotExist:
        return Decimal('0')

@login_required(login_url='/accounts/login/')
def monthly_rent(request, customer_id):
    customer_details = get_object_or_404(
        Bed.objects.select_related('unit', 'unit__hostel', 'customer'),
        customer=customer_id
    )

    # Validation: Check if required fees are set in bed details
    missing_fees = []
    if not customer_details.rent or customer_details.rent == 0:
        missing_fees.append("Base Rent")
    # interent fee can be 0
    # if not customer_details.internet_fee or customer_details.internet_fee == 0:
    #     missing_fees.append("Internet Fee")

    if not customer_details.utilities_fee or customer_details.utilities_fee == 0:
        missing_fees.append("Utilities Fee")
    
    if missing_fees:
        messages.error(request, f"Please add {', '.join(missing_fees)} in bed details. You cannot process without these fees.")
        return redirect('customer:customer_detail', customer_id)

    if request.method == "POST":
        month_input = request.POST.get("rent_month")
        if not month_input:
            messages.error(request, "Payment month is required.")
            return redirect(request.path)

        year, month = map(int, month_input.split("-"))
        
        # ===== COMPREHENSIVE RENT PAYMENT VALIDATION =====
        
        # 1. Registration Fee Validation
        registration_payment = HostelRevenue.objects.filter(
            title='registration_fee',
            customer=customer_details.customer
        ).first()
        
        if not registration_payment:
            messages.error(request, "Customer must pay registration fee before making rent payments.")
            return redirect(request.path)
        
        # 2. Get all existing rent payments for this customer
        existing_rent_payments = HostelRevenue.objects.filter(
            title='rent',
            customer=customer_details.customer
        ).order_by('year', 'month')
        
        # 3. Check for duplicate payment
        duplicate_payment = existing_rent_payments.filter(year=year, month=month).first()
        if duplicate_payment:
            messages.error(request, f"Payment for {year}-{month:02d} already exists. Please select a different month.")
            return redirect(request.path)
        
        # 4. First rent payment validation
        if not existing_rent_payments.exists():
            # This is the first rent payment - must be in registration month
            reg_year = registration_payment.year
            reg_month = registration_payment.month
            
            if year != reg_year or month != reg_month:
                messages.error(request, f"First rent payment must be in the registration month. Customer registered in {reg_year}-{reg_month:02d}. First rent payment should be for {reg_year}-{reg_month:02d}.")
                return redirect(request.path)
        else:
            # Not first payment - validate sequential payment
            last_payment = existing_rent_payments.last()
            last_year = last_payment.year
            last_month = last_payment.month
            
            # Calculate expected next month
            expected_month = last_month + 1
            expected_year = last_year
            if expected_month > 12:
                expected_month = 1
                expected_year += 1
            
            # Check if trying to pay for a month before the last payment
            if year < last_year or (year == last_year and month < last_month):
                messages.error(request, f"Cannot pay rent for a month before the last payment. Last payment was for {last_year}-{last_month:02d}.")
                return redirect(request.path)
            
            # Check if trying to skip months
            if year != expected_year or month != expected_month:
                messages.error(request, f"Cannot skip months. Last payment was for {last_year}-{last_month:02d}. Next payment should be for {expected_year}-{expected_month:02d}.")
                return redirect(request.path)
        
        # ===== END VALIDATION =====

        try:
            base_rent = Decimal(request.POST.get("rent", "0"))
            internet_fee = Decimal(request.POST.get("internet", "0"))
            utilities_fee = Decimal(request.POST.get("utilities", "0"))
            rent_discount_percent = Decimal(request.POST.get("rent_discount_percent", "0"))
            payment_type = request.POST.get("payment_type", "")
            
            # Handle collected_amount - it might be empty or contain invalid values
            collected_amount_str = request.POST.get("collected_amount", "").strip()
            if collected_amount_str and collected_amount_str != "" and collected_amount_str != "NaN":
                collected_amount = Decimal(collected_amount_str)
            else:
                collected_amount = Decimal("0")
            
            # Handle prepaid_amount - it might be empty for normal payments
            prepaid_amount_str = request.POST.get("prepaid_amount", "").strip()
            if prepaid_amount_str and prepaid_amount_str != "" and prepaid_amount_str != "NaN":
                prepaid_amount = Decimal(prepaid_amount_str)
            else:
                prepaid_amount = Decimal("0")
                
        except InvalidOperation as e:
            messages.error(request, f"Invalid numeric values in the form. Please check all amount fields. Error: {str(e)}")
            return redirect(request.path)

        rent_after_discount = base_rent * (Decimal(1) - rent_discount_percent / Decimal(100))
        total_amount = rent_after_discount + internet_fee + utilities_fee

        memo = request.POST.get("memo", "").strip()

        if rent_discount_percent > 0 and not memo:
            messages.error(request, "Memo is required when a discount is applied.")
            return redirect(request.path)

        # Compute only immediate previous month's prepaid/postpaid amounts
        previous_prepaid = Decimal('0')
        previous_postpaid = Decimal('0')
        
        prev_month = month - 1
        prev_year = year
        if prev_month <= 0:
            prev_month = 12
            prev_year = year - 1
        
        had_postpaid_last_month = False
        try:
            prev_revenue = HostelRevenue.objects.get(
                title='rent',
                customer=customer_details.customer,
                year=prev_year,
                month=prev_month
            )
            if prev_revenue.payment_type == 'prepaid' and prev_revenue.prepaid_amount:
                previous_prepaid = prev_revenue.prepaid_amount
            elif prev_revenue.payment_type == 'postpaid' and prev_revenue.prepaid_amount:
                previous_postpaid = prev_revenue.prepaid_amount
                had_postpaid_last_month = True
        except HostelRevenue.DoesNotExist:
            had_postpaid_last_month = False
        
        # Adjust total_amount for previous prepaid and add previous postpaid
        adjusted_total = total_amount - previous_prepaid + previous_postpaid
        
        # Restrict postpaid payment if customer had postpaid in previous month
        if had_postpaid_last_month and payment_type == 'postpaid':
            messages.error(request, "Customer cannot make postpaid payment this month as they had postpaid payment last month. Please pay the full amount including the previous postpaid amount.")
            return redirect(request.path)
        
        # Validate prepaid/postpaid logic
        if payment_type == 'prepaid':
            if collected_amount < adjusted_total:
                messages.error(request, f"For prepaid payment, collected amount must be greater than or equal to adjusted total rent amount (¥{adjusted_total}).")
                return redirect(request.path)
            expected_prepaid = collected_amount - adjusted_total
            if abs(prepaid_amount - expected_prepaid) > Decimal("0.01"):
                messages.error(request, f"Prepaid amount should be {expected_prepaid} (excess of collected amount over adjusted total rent).")
                return redirect(request.path)
        elif payment_type == 'postpaid':
            if collected_amount >= adjusted_total:
                messages.error(request, f"For postpaid payment, collected amount must be less than adjusted total rent amount (¥{adjusted_total}).")
                return redirect(request.path)
            expected_postpaid = adjusted_total - collected_amount
            if abs(prepaid_amount - expected_postpaid) > Decimal("0.01"):
                messages.error(request, f"Postpaid amount should be {expected_postpaid} (shortfall of adjusted total rent over collected amount).")
                return redirect(request.path)
        elif payment_type == '' and collected_amount == 0:
            # Normal payment - set collected_amount to adjusted_total (after prepaid deduction)
            collected_amount = adjusted_total
        elif payment_type == '' and abs(collected_amount - adjusted_total) > Decimal("0.01"):
            messages.error(request, f"For normal payment, collected amount must equal amount to collect (¥{adjusted_total}).")
            return redirect(request.path)

        revenue, created = HostelRevenue.objects.get_or_create(
            title="rent",
            customer=customer_details.customer,
            year=year,
            month=month,
            defaults={
                "rent": base_rent,
                "rent_discount_percent": rent_discount_percent,
                "rent_after_discount": rent_after_discount,
                "internet": internet_fee,
                "utilities": utilities_fee,
                "total_amount": total_amount,
                "payment_type": payment_type,
                "collected_amount": collected_amount,
                "prepaid_amount": prepaid_amount if payment_type else None,
                "memo": memo,
                "created_by": request.user,
                "updated_by": request.user,
            }
        )

        if not created:
            messages.warning(request, "Rent payment for this month already exists.")
        else:
            customer = revenue.customer
            if customer and customer.email:
                send_revenue_email(request, revenue, subject='Rent Payment Notification - Fishtail')

            messages.success(request, "Monthly rent payment recorded successfully.")
        return redirect("finance:revenues")

    # Check user privileges for editing fees
    can_edit_fees = request.user.has_perm('finance.change_hostelrevenue') or request.user.is_superuser

    # For GET requests, don't calculate previous prepaid until month is selected
    if request.method == "GET":
        previous_prepaid = Decimal('0')
        previous_postpaid = Decimal('0')
    else:
        # For POST requests, calculate based on selected month
        previous_prepaid = get_previous_prepaid_amount(customer_details.customer, year, month)
        
        # Get previous month's postpaid balance
        previous_postpaid = Decimal('0')
        prev_month = month - 1
        prev_year = year
        if prev_month <= 0:
            prev_month = 12
            prev_year = year - 1
        
        try:
            prev_revenue = HostelRevenue.objects.get(
                title='rent',
                customer=customer_details.customer,
                year=prev_year,
                month=prev_month
            )
            if prev_revenue.payment_type == 'postpaid' and prev_revenue.prepaid_amount:
                previous_postpaid = prev_revenue.prepaid_amount
        except HostelRevenue.DoesNotExist:
            previous_postpaid = Decimal('0')
    
    # Check if customer had postpaid payment in previous month (to restrict postpaid in current month)
    # This will be handled dynamically via AJAX when month is selected
    had_postpaid_last_month = False

    # Debug information will be shown dynamically when month is selected
    
    # Get payment history for context
    registration_payment = HostelRevenue.objects.filter(
        title='registration_fee',
        customer=customer_details.customer
    ).first()
    
    existing_rent_payments = HostelRevenue.objects.filter(
        title='rent',
        customer=customer_details.customer
    ).order_by('year', 'month')
    
    # Get last rent payment
    last_rent_payment = existing_rent_payments.last() if existing_rent_payments.exists() else None

    return render(request, 'finance/monthly_rent.html', {
        'customer_details': customer_details,
        'can_edit_fees': can_edit_fees,
        'previous_prepaid_amount': previous_prepaid,
        'previous_postpaid_amount': previous_postpaid,
        'had_postpaid_last_month': had_postpaid_last_month,
        'registration_payment': registration_payment,
        'existing_rent_payments': existing_rent_payments,
        'last_rent_payment': last_rent_payment
    })


@login_required(login_url='/accounts/login/')
def get_prepaid_amount_for_month(request, customer_id):
    """AJAX endpoint to get previous month's prepaid/postpaid for a selected month"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            year = int(data.get('year'))
            month = int(data.get('month'))
            
            # Calculate previous month
            prev_month = month - 1
            prev_year = year
            if prev_month <= 0:
                prev_month = 12
                prev_year = year - 1
            
            # Get customer
            customer = get_object_or_404(Bed.objects.select_related('customer'), customer=customer_id).customer
            
            previous_prepaid = Decimal('0')
            previous_postpaid = Decimal('0')
            had_postpaid_last_month = False
            try:
                prev_revenue = HostelRevenue.objects.get(
                    title='rent',
                    customer=customer,
                    year=prev_year,
                    month=prev_month
                )
                if prev_revenue.payment_type == 'prepaid' and prev_revenue.prepaid_amount:
                    previous_prepaid = prev_revenue.prepaid_amount
                elif prev_revenue.payment_type == 'postpaid' and prev_revenue.prepaid_amount:
                    previous_postpaid = prev_revenue.prepaid_amount
                    had_postpaid_last_month = True
            except HostelRevenue.DoesNotExist:
                pass
            
            return JsonResponse({
                'success': True,
                'previous_prepaid': float(previous_prepaid),
                'previous_postpaid': float(previous_postpaid),
                'had_postpaid_last_month': had_postpaid_last_month,
                'previous_month': f"{prev_year}-{prev_month:02d}"
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required(login_url='/accounts/login/')
def validate_rent_month(request, customer_id):
    """AJAX endpoint to validate rent payment month selection"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            year = int(data.get('year'))
            month = int(data.get('month'))
            
            # Get customer
            customer = get_object_or_404(Bed.objects.select_related('customer'), customer=customer_id).customer
            
            # 1. Registration Fee Validation
            registration_payment = HostelRevenue.objects.filter(
                title='registration_fee',
                customer=customer
            ).first()
            
            if not registration_payment:
                return JsonResponse({
                    'success': False,
                    'error': 'Customer must pay registration fee before making rent payments.',
                    'error_type': 'registration_required'
                })
            
            # 2. Get all existing rent payments for this customer
            existing_rent_payments = HostelRevenue.objects.filter(
                title='rent',
                customer=customer
            ).order_by('year', 'month')
            
            # 3. Check for duplicate payment
            duplicate_payment = existing_rent_payments.filter(year=year, month=month).first()
            if duplicate_payment:
                return JsonResponse({
                    'success': False,
                    'error': f'Payment for {year}-{month:02d} already exists. Please select a different month.',
                    'error_type': 'duplicate_payment'
                })
            
            # 4. First rent payment validation
            if not existing_rent_payments.exists():
                # This is the first rent payment - must be in registration month
                reg_year = registration_payment.year
                reg_month = registration_payment.month
                
                if year != reg_year or month != reg_month:
                    return JsonResponse({
                        'success': False,
                        'error': f'First rent payment must be in the registration month. Customer registered in {reg_year}-{reg_month:02d}. First rent payment should be for {reg_year}-{reg_month:02d}.',
                        'error_type': 'wrong_first_month',
                        'suggested_month': f'{reg_year}-{reg_month:02d}'
                    })
            else:
                # Not first payment - validate sequential payment
                last_payment = existing_rent_payments.last()
                last_year = last_payment.year
                last_month = last_payment.month
                
                # Calculate expected next month
                expected_month = last_month + 1
                expected_year = last_year
                if expected_month > 12:
                    expected_month = 1
                    expected_year += 1
                
                # Check if trying to pay for a month before the last payment
                if year < last_year or (year == last_year and month < last_month):
                    return JsonResponse({
                        'success': False,
                        'error': f'Cannot pay rent for a month before the last payment. Last payment was for {last_year}-{last_month:02d}.',
                        'error_type': 'backward_payment',
                        'last_payment': f'{last_year}-{last_month:02d}'
                    })
                
                # Check if trying to skip months
                if year != expected_year or month != expected_month:
                    return JsonResponse({
                        'success': False,
                        'error': f'Cannot skip months. Last payment was for {last_year}-{last_month:02d}. Next payment should be for {expected_year}-{expected_month:02d}.',
                        'error_type': 'skip_months',
                        'last_payment': f'{last_year}-{last_month:02d}',
                        'suggested_month': f'{expected_year}-{expected_month:02d}'
                    })
            
            # If we get here, validation passed
            return JsonResponse({
                'success': True,
                'message': 'Month selection is valid.'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Validation error: {str(e)}',
                'error_type': 'system_error'
            })
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required(login_url='/accounts/login/')
def registration_fee(request, customer_id):
    customer_details = get_object_or_404(
        Bed.objects.select_related('unit', 'unit__hostel', 'customer'),
        customer=customer_id
    )

    # Validation: Check if required fees are set in hostel details
    missing_fees = []
    if not customer_details.unit.hostel.deposit_fee or customer_details.unit.hostel.deposit_fee == 0:
        missing_fees.append("Deposit Fee")
    if not customer_details.unit.hostel.initial_fee or customer_details.unit.hostel.initial_fee == 0:
        missing_fees.append("Initial Fee")
    
    if missing_fees:
        messages.error(request, f"Please add {', '.join(missing_fees)} in hostel details. You cannot process without these fees.")
        return redirect('customer:customer_detail', customer_id)

    if request.method == "POST":
        month_input = request.POST.get("reg_month")
        if not month_input:
            messages.error(request, "Payment month is required.")
            return redirect(request.path)

        try:
            year, month = map(int, month_input.split("-"))
        except ValueError:
            messages.error(request, "Invalid month format.")
            return redirect(request.path)

        try:
            deposit = Decimal(request.POST.get("deposit", "0"))
            deposit_discount = Decimal(request.POST.get("deposit_discount_percent", "0"))
            initial = Decimal(request.POST.get("initial_fee", "0"))
            initial_discount = Decimal(request.POST.get("initial_fee_discount_percent", "0"))
        except InvalidOperation:
            messages.error(request, "Invalid numeric values.")
            return redirect(request.path)

        deposit_after = deposit * (Decimal(1) - deposit_discount / Decimal(100))
        initial_after = initial * (Decimal(1) - initial_discount / Decimal(100))
        total = deposit_after + initial_after

        memo = request.POST.get("memo", "").strip()
        if (deposit_discount > 0 or initial_discount > 0) and not memo:
            messages.error(request, "Memo is required when a discount is applied.")
            return redirect(request.path)

        # Prevent duplicate
        revenue, created = HostelRevenue.objects.get_or_create(
            title="registration_fee",
            customer=customer_details.customer,
            year=year,
            month=month,
            defaults={
                "deposit": deposit,
                "deposit_discount_percent": deposit_discount,
                "deposit_after_discount": deposit_after,
                "initial_fee": initial,
                "initial_fee_discount_percent": initial_discount,
                "initial_fee_after_discount": initial_after,
                "total_amount": total,
                "memo": memo,
                "created_by": request.user,
                "updated_by": request.user,
            }
        )

        if not created:
            messages.warning(request, "Registration fee for this month already exists.")
        else:
            customer = revenue.customer
            if customer and customer.email:
                send_revenue_email(request, revenue, subject='Registration Fee Notification - Fishtail')

            messages.success(request, "Registration fee payment recorded successfully.")

        return redirect("finance:revenues")

    # Check user privileges for editing fees
    can_edit_fees = request.user.has_perm('finance.change_hostelrevenue') or request.user.is_superuser

    return render(request, 'finance/registration_fee.html', {
        'customer_details': customer_details,
        'can_edit_fees': can_edit_fees
    })


@login_required(login_url='/accounts/login/')
def notification(request):
    search_name = request.GET.get('name', '').strip().lower()
    defaulters = get_rent_defaulters()

    if search_name:
        defaulters = [
            d for d in defaulters
            if search_name in d['customer'].name.lower()
        ]

    # Handle Excel download
    if 'download' in request.GET:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Unpaid Rent'

        # Header
        sheet.append(['Customer Name', 'Stay Type', 'Assigned Date', 'Released/End Date', 'Unpaid Months'])

        for entry in defaulters:
            unpaid_str = ", ".join(f"{y}-{m:02d}" for y, m in entry['unpaid_months'])
            sheet.append([
                entry['customer'].name,
                entry['type'].capitalize(),
                entry['assigned_date'],
                entry['end_date'],
                unpaid_str
            ])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="unpaid_rent.xlsx"'
        workbook.save(response)
        return response

    return render(request, 'finance/notification.html', {
        'defaulters': defaulters,
        'search_name': request.GET.get('name', ''),
        'today': date.today()
        
    })


@login_required(login_url='/accounts/login/')
@permission_required(('finance.view_hostelexpense', 'finance.view_utilityexpense'), raise_exception=True)
def expenses(request):
    # Get filters from request
    year_month = request.GET.get('year_month')
    status = request.GET.get('status')
    expense_type = request.GET.get('expense_type')
    hostel_filter = request.GET.get('hostel')
    export = request.GET.get('export')

    # Get both types of expenses
    hostel_expenses = HostelExpense.objects.select_related('hostel', 'created_by', 'updated_by', 'approved_by')
    utility_expenses = UtilityExpense.objects.select_related('hostel', 'paid_by', 'approved_by')

    # Default to current month if no year_month filter is provided
    if not year_month and not status and not hostel_filter and not export:
        current_date = date.today()
        year_month = current_date.strftime('%Y-%m')

    # Filter by year and month (if specified)
    if year_month:
        try:
            filter_year, filter_month = map(int, year_month.split('-'))
            
            # For hostel expenses: filter by purchased_date year and month
            hostel_expenses = hostel_expenses.filter(
                purchased_date__year=filter_year,
                purchased_date__month=filter_month
            )
            
            # For utility expenses: filter by paid_date year and month
            utility_expenses = utility_expenses.filter(
                paid_date__year=filter_year,
                paid_date__month=filter_month
            )
        except (ValueError, TypeError):
            # Invalid year_month format, show current month
            current_date = date.today()
            filter_year, filter_month = current_date.year, current_date.month
            hostel_expenses = hostel_expenses.filter(
                purchased_date__year=filter_year,
                purchased_date__month=filter_month
            )
            utility_expenses = utility_expenses.filter(
                paid_date__year=filter_year,
                paid_date__month=filter_month
            )

    # Filter by status
    if status in ['approved', 'pending', 'rejected']:
        hostel_expenses = hostel_expenses.filter(status=status)
        utility_expenses = utility_expenses.filter(approval_status=status.upper())

    # Filter by expense type
    if expense_type == 'hostel':
        utility_expenses = utility_expenses.none()
    elif expense_type == 'utility':
        hostel_expenses = hostel_expenses.none()

    # Filter by hostel
    if hostel_filter:
        hostel_expenses = hostel_expenses.filter(hostel__name__icontains=hostel_filter)
        utility_expenses = utility_expenses.filter(hostel__name__icontains=hostel_filter)

    # Combine and sort expenses
    combined_expenses = []
    
    # Add hostel expenses
    #print(f"DEBUG: Found {hostel_expenses.count()} hostel expenses")
    for expense in hostel_expenses:
        #print(f"DEBUG: Processing hostel expense ID {expense.id}")
        combined_expenses.append({
            'id': expense.id,
            'type': 'hostel',
            'transaction_code': expense.transaction_code,
            'date': expense.purchased_date,
            'date_display': expense.purchased_date.strftime('%b %d, %Y') if expense.purchased_date else 'N/A',
            'hostel': expense.hostel.name if expense.hostel else "ALL",
            'purchased_by': expense.purchased_by,
            'memo': expense.memo,
            'amount': expense.amount,
            'status': expense.status,
            'approved_by': expense.approved_by,
            'created_by': expense.created_by,
            'created_at': expense.created_at,
            'updated_by': expense.updated_by,
            'updated_at': expense.updated_at,
        })
    
    # Add utility expenses
    #print(f"DEBUG: Found {utility_expenses.count()} utility expenses")
    for expense in utility_expenses:
        #print(f"DEBUG: Processing utility expense ID {expense.id}")
        # Create a date object for sorting (first day of billing month)
        billing_date = date(expense.billing_year, expense.billing_month, 1) if expense.billing_year and expense.billing_month else None
        
        combined_expenses.append({
            'id': expense.id,
            'type': expense.expense_type,
            'transaction_code': f"UTIL-{expense.id:06d}",
            'date': billing_date,  # For sorting
            'date_display': f"{expense.paid_date.strftime('%b %d, %Y') if expense.paid_date else 'N/A'}<br><small class='text-muted mb-0'>(Bill of {['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec'][expense.billing_month-1]} {expense.billing_year})</small>" if expense.paid_date and expense.billing_year and expense.billing_month else (expense.paid_date.strftime('%b %d, %Y') if expense.paid_date else 'N/A'),
            'hostel': expense.hostel.name,
            'purchased_by': expense.paid_by.first_name if expense.paid_by else "N/A",
            'memo': expense.description,
            'amount': expense.amount,
            'status': expense.approval_status.lower(),
            'approved_by': expense.approved_by,
            'created_by': expense.created_by,
            'created_at': expense.created_at,
            'updated_by': expense.updated_by,
            'updated_at': expense.updated_at,
        })

    # Sort by date (newest first)
    combined_expenses.sort(key=lambda x: x['date'] if x['date'] else date.min, reverse=True)
    #print(f"DEBUG: Total combined expenses: {len(combined_expenses)}")

    # Calculate totals
    total_amount = sum([
        expense['amount'] if expense['amount'] is not None else Decimal('0')
        for expense in combined_expenses
    ], Decimal('0'))

    # Paginate combined expenses
    page_number = request.GET.get('page')
    paginator = Paginator(combined_expenses, 25)
    page_obj = paginator.get_page(page_number)
    expenses_page = page_obj.object_list

    page_total_amount = sum([
        expense['amount'] if expense['amount'] is not None else Decimal('0')
        for expense in expenses_page
    ], Decimal('0'))

    # Preserve current filters in pagination links
    query_params = request.GET.copy()
    if 'page' in query_params:
        query_params.pop('page')
    query_string = query_params.urlencode()

    # Export to Excel
    if export == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "All Expenses"

        ws.append([
            "Type", "ID", "Date", "Hostel", "Purchased By", "Approved By", "Amount", "Status", "Memo",
            "Created By", "Created At", "Updated By", "Updated At"
        ])

        for e in combined_expenses:
            approved_name = "-"
            if e['approved_by']:
                approved_name = e['approved_by'].first_name or e['approved_by'].email

            ws.append([
                e['type'].title(),
                e['transaction_code'],
                e['date_display'],
                e['hostel'],
                e['purchased_by'],
                approved_name,
                float(e['amount']),
                e['status'],
                e['memo'] or "-",
                str(e['created_by']) if e['created_by'] else "-",
                e['created_at'].strftime('%Y-%m-%d %H:%M:%S') if e['created_at'] else "-",
                str(e['updated_by']) if e['updated_by'] else "-",
                e['updated_at'].strftime('%Y-%m-%d %H:%M:%S') if e['updated_at'] else "-",
            ])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=all_expenses.xlsx'
        wb.save(response)
        return response

    # Get all hostels for the filter dropdown
    from hostel.models import Hostel
    all_hostels = Hostel.objects.all().order_by('name')
    
    # Format year_month for display
    display_month = ""
    if year_month:
        try:
            filter_year, filter_month = map(int, year_month.split('-'))
            from datetime import datetime
            display_month = datetime(filter_year, filter_month, 1).strftime('%B %Y')
        except (ValueError, TypeError):
            display_month = "Current Month"
    
    #print(combined_expenses)
    return render(request, 'finance/expenses_dashboard.html', {
        'expenses': expenses_page,
        'year_month': year_month,
        'display_month': display_month,
        'status': status,
        'expense_type': expense_type,
        'hostel_filter': hostel_filter,
        'all_hostels': all_hostels,
        'page_obj': page_obj,
        'paginator': paginator,
        'total_records': paginator.count,
        'total_amount': total_amount,
        'page_total_amount': page_total_amount,
        'query_string': query_string,
    })

@login_required(login_url='/accounts/login/')
@permission_required('finance.add_hostelexpense', raise_exception=True)
def hostel_expense_create(request):
    if request.method == 'POST':
        form = HostelExpenseForm(request.POST)
        if form.is_valid():
            expense = form.save(commit=False)
            expense.created_by = request.user
            expense.updated_by = request.user
            expense.save()
            messages.success(request, "Expense created successfully.")
            return redirect('finance:expenses')
    else:
        form = HostelExpenseForm()
    return render(request, 'finance/hostel_expense_form.html', {'form': form})


@login_required(login_url='/accounts/login/')
@permission_required('finance.change_hostelexpense', raise_exception=True)
def hostel_expense_edit(request, pk):
    expense = get_object_or_404(HostelExpense, pk=pk)

    if expense.status == 'approved':
        messages.warning(request, "Approved expenses cannot be edited.")
        return redirect('finance:expenses')

    if request.method == 'POST':
        form = HostelExpenseForm(request.POST, instance=expense)
        if form.is_valid():
            expense = form.save(commit=False)
            expense.updated_by = request.user
            expense.save()
            messages.success(request, "Expense updated successfully.")
            return redirect('finance:expenses')
    else:
        form = HostelExpenseForm(instance=expense)

    return render(request, 'finance/hostel_expense_form.html', {'form': form})


@login_required(login_url='/accounts/login/')
@user_passes_test(lambda u: u.is_superuser)
def hostel_expense_detail(request, pk):
    expense = get_object_or_404(HostelExpense, pk=pk)

    if request.method == 'POST':
        new_status = request.POST.get('status')
        if new_status in dict(HostelExpense.STATUS_CHOICES):
            expense.status = new_status
            if new_status == 'approved':
                expense.updated_by = request.user
                expense.approved_by = request.user
            else:
                expense.approved_by = None
            expense.save()
            messages.success(request, "Status updated successfully.")
            return redirect('finance:expenses')

    return render(request, 'finance/hostel_expense_detail.html', {'expense': expense})


# Utility Expense Views
@login_required(login_url='/accounts/login/')
@permission_required('finance.add_utilityexpense', raise_exception=True)
def utility_expense_create(request):
    if request.method == 'POST':
        form = UtilityExpenseForm(request.POST, request.FILES)
        if form.is_valid():
            expense = form.save(commit=False)
            expense.paid_by = request.user
            expense.created_by = request.user
            expense.updated_by = request.user
            expense.save()
            messages.success(request, "Utility expense created successfully.")
            return redirect('finance:expenses')
    else:
        form = UtilityExpenseForm()
    return render(request, 'finance/utility_expense_form.html', {'form': form})


@login_required(login_url='/accounts/login/')
@permission_required('finance.change_utilityexpense', raise_exception=True)
def utility_expense_edit(request, pk):
    expense = get_object_or_404(UtilityExpense, pk=pk)

    if expense.approval_status == 'APPROVED':
        messages.warning(request, "Approved utility expenses cannot be edited.")
        return redirect('finance:expenses')

    if request.method == 'POST':
        form = UtilityExpenseForm(request.POST, request.FILES, instance=expense)
        if form.is_valid():
            expense = form.save(commit=False)
            expense.updated_by = request.user
            expense.save()
            messages.success(request, "Utility expense updated successfully.")
            return redirect('finance:expenses')
    else:
        form = UtilityExpenseForm(instance=expense)

    return render(request, 'finance/utility_expense_form.html', {'form': form})


@login_required(login_url='/accounts/login/')
@user_passes_test(lambda u: u.is_superuser)
def utility_expense_detail(request, pk):
    expense = get_object_or_404(UtilityExpense, pk=pk)

    if request.method == 'POST':
        new_status = request.POST.get('approval_status')
        if new_status in dict(UtilityExpense.ApprovalStatus.choices):
            expense.approval_status = new_status
            if new_status == 'APPROVED':
                expense.approved_by = request.user
            else:
                expense.approved_by = None
            expense.updated_by = request.user
            expense.save()
            messages.success(request, "Status updated successfully.")
            return redirect('finance:expenses')

    return render(request, 'finance/utility_expense_detail.html', {'expense': expense})
